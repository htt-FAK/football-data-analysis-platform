"""导出 Football-Data 清洗前后对比 Excel

生成 3 个 Sheet：
1. 清洗前-原始数据：爬虫返回的 raw（CSV 原始字段）
2. 清洗后-干净数据：数据库 Match 表（标准化字段 + 实体解析）
3. 清洗前后对比：关键字段对照（Date→match_date, HomeTeam→home_team 等）

用法: python export_football_data.py
输出: <项目根目录>/export/football_data/football_data_<ts>.xlsx
"""
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import SessionLocal
from app.crawlers.football_data import FootballDataCrawler
from app.models.match import Match
from app.models.team import Team
from app.models.season import Season

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = PROJECT_ROOT / "export" / "football_data"
# (league_code, season_code, league_name, season_name)
SEASONS = [
    ("E0", "2425", "英超", "2024-2025"),
    ("E0", "2526", "英超", "2025-2026"),
]

# 清洗前 raw 关键字段（Football-Data CSV 原始列名）
RAW_COLS = ["Date", "Time", "HomeTeam", "AwayTeam", "FTHG", "FTAG", "FTR",
            "HS", "AS", "HST", "AST", "HY", "AY", "HR", "AR",
            "HC", "AC", "HF", "AF", "Referee"]

# 清洗后字段（数据库标准化后）
CLEAN_COLS = ["id", "season", "match_date", "status", "home_team", "away_team",
              "home_score", "away_score", "source_id", "version"]

# 对比映射：原始字段 → 清洗后字段
COMPARE_MAP = [
    ("Date", "match_date", "日期解析 dd/mm/yyyy → ISO datetime"),
    ("Time", "match_date", "时间合并到 match_date"),
    ("HomeTeam", "home_team", "球队名标准化（英文别名 → 规范名）"),
    ("AwayTeam", "away_team", "球队名标准化"),
    ("FTHG", "home_score", "主队进球（字段名标准化）"),
    ("FTAG", "away_score", "客队进球"),
    ("FTR", "status", "结果码 H/D/A → status=finished（有比分推断）"),
]


def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def write_df(ws, df: pd.DataFrame):
    for j, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=col_name)
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    style_header(ws)
    auto_width(ws)


def main():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(str(EXPORT_DIR), f"football_data_{ts}.xlsx")

    db = SessionLocal()
    crawler = FootballDataCrawler()

    all_raw = []
    all_clean = []

    for league_code, season_code, league_name, season_name in SEASONS:
        print(f"采集 {season_name} raw ...")
        raw = crawler.crawl(target="matches", league=league_code, season=season_code)
        print(f"  raw: {len(raw)} 场")
        for r in raw:
            row = {c: r.get(c) for c in RAW_COLS}
            row["season"] = season_name
            all_raw.append(row)

        # 清洗后：从数据库读
        season_obj = db.query(Season).filter(Season.name == season_name).first()
        if season_obj:
            matches = db.query(Match).filter(Match.season_id == season_obj.id).order_by(Match.id).all()
            print(f"  db: {len(matches)} 场")
            for m in matches:
                home = db.query(Team).get(m.home_team_id) if m.home_team_id else None
                away = db.query(Team).get(m.away_team_id) if m.away_team_id else None
                all_clean.append({
                    "id": m.id, "season": season_name,
                    "match_date": m.match_date, "status": m.status,
                    "home_team": home.name if home else None,
                    "away_team": away.name if away else None,
                    "home_score": m.home_score, "away_score": m.away_score,
                    "source_id": m.source_id, "version": m.version,
                })

    db.close()

    # 写 Excel
    wb = Workbook()

    # Sheet 1: 清洗前
    ws1 = wb.active
    ws1.title = "清洗前-原始数据"
    write_df(ws1, pd.DataFrame(all_raw))

    # Sheet 2: 清洗后
    ws2 = wb.create_sheet("清洗后-干净数据")
    write_df(ws2, pd.DataFrame(all_clean))

    # Sheet 3: 对比（取每赛季第 1 场做字段级对照）
    ws3 = wb.create_sheet("清洗前后对比")
    compare_rows = []
    for season_name in [s[3] for s in SEASONS]:
        raw_sample = next((r for r in all_raw if r["season"] == season_name), None)
        clean_sample = next((c for c in all_clean if c["season"] == season_name), None)
        if not raw_sample or not clean_sample:
            continue
        for raw_field, clean_field, note in COMPARE_MAP:
            compare_rows.append({
                "赛季": season_name,
                "原始字段": raw_field,
                "原始值": raw_sample.get(raw_field),
                "清洗后字段": clean_field,
                "清洗后值": clean_sample.get(clean_field),
                "处理说明": note,
            })
    write_df(ws3, pd.DataFrame(compare_rows))

    # Sheet 4: 概览
    ws4 = wb.create_sheet("导出概览")
    overview = [
        {"项目": "导出时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"项目": "数据源", "值": "Football-Data.co.uk"},
        {"项目": "赛季", "值": " + ".join(s[3] for s in SEASONS)},
        {"项目": "清洗前记录数", "值": len(all_raw)},
        {"项目": "清洗后记录数", "值": len(all_clean)},
        {"项目": "清洗字段数", "值": len(RAW_COLS)},
        {"项目": "标准化字段数", "值": len(CLEAN_COLS)},
    ]
    write_df(ws4, pd.DataFrame(overview))

    wb.save(filepath)
    print(f"\n导出完成: {filepath}")
    print(f"清洗前: {len(all_raw)} 条, 清洗后: {len(all_clean)} 条")
    print(f"共 4 个 Sheet: 清洗前 / 清洗后 / 对比 / 概览")


if __name__ == "__main__":
    main()
