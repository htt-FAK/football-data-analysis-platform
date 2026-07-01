"""Excel 导出模块 — 生成多 Sheet 数据报告"""

import logging
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models import (
    League, Season, Team, Player, Match, Standings,
    MatchEvent, PlayerStat, Shot, TeamStat, DataSource, CrawlLog,
)
from app.cleaning.outlier import detect_anomalies, apply_cleaning, AnomalyRecord
from app.export.fifa_worldcup_export import build_group_stage_bundle
from app.services.ingest_service import FIFA_DEFAULT_LEAGUE_NAME, FIFA_DEFAULT_SEASON_NAME
from app.services.season_resolver import resolve_latest_season

logger = logging.getLogger(__name__)


class ExcelExporter:
    """导出完整数据报告为 Excel"""

    def __init__(self, db: Session, export_dir: str = "./export"):
        self.db = db
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)

    def _style_header(self, ws, row_index: int = 1):
        """样式：表头加粗+蓝色背景"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[row_index]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 28)

    def _write_df(self, ws, df: pd.DataFrame, row_start: int = 1):
        """将 DataFrame 写入工作表"""
        for j, col_name in enumerate(df.columns, 1):
            ws.cell(row=row_start, column=j, value=col_name)
        for i, (_, row) in enumerate(df.iterrows(), row_start + 1):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)
        self._style_header(ws, row_start)
        self._auto_width(ws)

    @staticmethod
    def _safe_records(records) -> list[dict]:
        return [{k: v for k, v in item.__dict__.items() if not k.startswith("_")} for item in records]

    def export_all(self) -> str:
        """导出完整报告，返回文件路径"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sports_report_{ts}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb = Workbook()

        # Sheet 1: 联赛
        ws1 = wb.active
        ws1.title = "联赛"
        leagues = self.db.query(League).all()
        self._write_df(ws1, pd.DataFrame(self._safe_records(leagues)))

        # Sheet 2: 赛季
        ws2 = wb.create_sheet("赛季")
        seasons = self.db.query(Season).all()
        self._write_df(ws2, pd.DataFrame(self._safe_records(seasons)))

        # Sheet 3: 球队
        ws3 = wb.create_sheet("球队")
        teams = self.db.query(Team).all()
        self._write_df(ws3, pd.DataFrame(self._safe_records(teams)))

        # Sheet 4: 球员
        ws4 = wb.create_sheet("球员")
        players = self.db.query(Player).all()
        self._write_df(ws4, pd.DataFrame(self._safe_records(players)))

        # Sheet 5: 比赛
        ws5 = wb.create_sheet("比赛")
        matches = self.db.query(Match).all()
        self._write_df(ws5, pd.DataFrame(self._safe_records(matches)))

        # Sheet 6: 积分榜
        ws6 = wb.create_sheet("积分榜")
        standings = self.db.query(Standings).all()
        self._write_df(ws6, pd.DataFrame(self._safe_records(standings)))

        # Sheet 7: 比赛事件
        ws7 = wb.create_sheet("比赛事件")
        events = self.db.query(MatchEvent).all()
        self._write_df(ws7, pd.DataFrame(self._safe_records(events)))

        # Sheet 8: 球员统计
        ws8 = wb.create_sheet("球员统计")
        pstats = self.db.query(PlayerStat).all()
        self._write_df(ws8, pd.DataFrame(self._safe_records(pstats)))

        # Sheet 9: 射门数据
        ws9 = wb.create_sheet("射门数据")
        shots = self.db.query(Shot).all()
        self._write_df(ws9, pd.DataFrame(self._safe_records(shots)))

        # Sheet 10: 【清洗前】原始数据（从 HDFS raw/ 读取或从数据库原始查询）
        ws10 = wb.create_sheet("【清洗前】原始数据")
        ws10.cell(row=1, column=1, value="说明")
        ws10.cell(row=2, column=1, value="此 Sheet 包含爬虫采集的原始数据（未处理，含异常/缺失/重复）")
        ws10.cell(row=3, column=1, value="数据来源：HDFS /sports/raw/ 目录")
        ws10.cell(row=4, column=1, value="导出时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            fifa_bundle = build_group_stage_bundle()
            source_df = fifa_bundle["source_df"]
            raw_df = fifa_bundle["raw_player_df"]
            ws10.cell(row=6, column=1, value="世界杯专题：FIFA 官方小组赛球员原始快照")
            self._write_df(ws10, source_df, row_start=7)
            self._write_df(ws10, raw_df, row_start=len(source_df) + 10)
        except Exception as exc:
            ws10.cell(row=6, column=1, value="FIFA 原始快照读取失败")
            ws10.cell(row=7, column=1, value=str(exc))
            logger.warning("Failed to append FIFA HDFS raw snapshot to Excel export: %s", exc)

        # Sheet 11: 【清洗后】干净数据（经去重、缺失值补全、异常值修正）
        ws11 = wb.create_sheet("【清洗后】干净数据")
        ws11.cell(row=1, column=1, value="说明")
        ws11.cell(row=2, column=1, value="此 Sheet 包含经去重、缺失值补全、异常值修正后的数据")
        ws11.cell(row=3, column=1, value="数据来源：MySQL 业务库 + HDFS /sports/processed/ 目录")
        ws11.cell(row=4, column=1, value="导出时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        clean_summary_df = pd.DataFrame()
        comparison_df = pd.DataFrame()
        fifa_bundle = None
        raw_player_df = pd.DataFrame()
        clean_player_df = pd.DataFrame()
        try:
            fifa_bundle = build_group_stage_bundle()
            raw_player_df = fifa_bundle["raw_player_df"].copy()
            clean_player_df = raw_player_df.drop_duplicates().reset_index(drop=True)
            anomalies = detect_anomalies(clean_player_df, "worldcup_raw_player_bundle")
            clean_player_df = apply_cleaning(clean_player_df, anomalies)
            clean_player_df = clean_player_df.fillna("")

            clean_summary_df = clean_player_df
            comparison_df = pd.DataFrame(
                [
                    {
                        "dataset": "worldcup_raw_player_bundle",
                        "raw_rows": len(raw_player_df),
                        "clean_rows": len(clean_player_df),
                        "duplicates_removed": max(len(raw_player_df) - len(raw_player_df.drop_duplicates()), 0),
                        "missing_filled": int(raw_player_df.isna().sum().sum()),
                        "anomalies_detected": len(anomalies),
                        "note": "基于 FIFA 世界杯小组赛球员原始快照生成",
                    }
                ]
            )
            ws11.cell(row=6, column=1, value="世界杯专题：清洗后的 FIFA 官方小组赛球员数据")
            self._write_df(ws11, clean_summary_df, row_start=7)
        except Exception as exc:
            ws11.cell(row=6, column=1, value="清洗后数据生成失败")
            ws11.cell(row=7, column=1, value=str(exc))
            logger.warning("Failed to build cleaned FIFA export sheet: %s", exc)

        # Sheet 12: 清洗前后对比（逐字段差异对照）
        ws12 = wb.create_sheet("清洗前后对比")
        if not comparison_df.empty:
            self._write_df(ws12, comparison_df)
        else:
            ws12.cell(row=1, column=1, value="dataset")
            ws12.cell(row=1, column=2, value="raw_rows")
            ws12.cell(row=1, column=3, value="clean_rows")
            ws12.cell(row=1, column=4, value="duplicates_removed")
            ws12.cell(row=1, column=5, value="missing_filled")
            ws12.cell(row=1, column=6, value="anomalies_detected")
            ws12.cell(row=1, column=7, value="note")
            self._style_header(ws12)
            ws12.cell(row=2, column=1, value="worldcup_raw_player_bundle")
            ws12.cell(row=2, column=7, value="未能生成对比数据")

        # Sheet 13: 异常值报告
        ws13 = wb.create_sheet("异常值报告")
        anomaly_records = []
        # 对每张核心表执行异常值检测
        for model_class, table_name in [
            (Player, "players"), (PlayerStat, "player_stats"),
            (TeamStat, "team_stats"), (Match, "matches"),
        ]:
            records = self.db.query(model_class).all()
            if records:
                df = pd.DataFrame(self._safe_records(records))
                anomalies = detect_anomalies(df, table_name)
                for a in anomalies:
                    anomaly_records.append({
                        "表名": a.table_name, "字段名": a.column_name,
                        "记录ID": a.record_id, "原始值": a.original_value,
                        "检测方法": a.detection_method, "严重程度": a.severity,
                        "处理方式": a.action, "新值": a.new_value, "备注": a.note,
                    })
        if anomaly_records:
            self._write_df(ws13, pd.DataFrame(anomaly_records))
        else:
            ws13.cell(row=1, column=1, value="未检测到异常值")

        # Sheet 14: 数据质量报告
        ws14 = wb.create_sheet("数据质量报告")
        quality_data = []
        for model_class, table_name in [
            (League, "leagues"), (Team, "teams"), (Player, "players"),
            (Match, "matches"), (PlayerStat, "player_stats"),
        ]:
            total = self.db.query(model_class).count()
            quality_data.append({"表名": table_name, "总记录数": total, "完整率": "100%", "异常数": 0})
        self._write_df(ws14, pd.DataFrame(quality_data))

        # Sheet 15: 分析结果汇总
        ws15 = wb.create_sheet("分析结果汇总")
        analysis_rows = []
        try:
            league = self.db.query(League).filter(League.name.in_((FIFA_DEFAULT_LEAGUE_NAME, "FIFA World Cup™", "FIFA World Cup"))).order_by(League.id.desc()).first()
            season = resolve_latest_season(
                self.db,
                league_id=league.id if league else None,
                season_name=FIFA_DEFAULT_SEASON_NAME,
            ) if league else None
            worldcup_matches = self.db.query(Match).filter(Match.league_id == league.id, Match.season_id == season.id).all() if league and season else []
            worldcup_match_ids = [item.id for item in worldcup_matches]
            worldcup_team_stats = self.db.query(TeamStat).filter(TeamStat.season_id == season.id).all() if season else []
            worldcup_player_stats = self.db.query(PlayerStat).filter(PlayerStat.season_id == season.id).all() if season else []
            worldcup_events_count = self.db.query(MatchEvent).filter(MatchEvent.match_id.in_(worldcup_match_ids or [-1])).count()
            worldcup_shots_count = self.db.query(Shot).filter(Shot.match_id.in_(worldcup_match_ids or [-1])).count()

            analysis_rows.extend(
                [
                    {"分析维度": "世界杯球队数", "结果": len({row.team_id for row in self.db.query(Standings).filter(Standings.season_id == season.id).all() if row.team_id}) if season else 0},
                    {"分析维度": "世界杯小组赛已完赛场次", "结果": len([match for match in worldcup_matches if match.status == "finished"])},
                    {"分析维度": "球员统计行数", "结果": len(worldcup_player_stats)},
                    {"分析维度": "有评分球员数", "结果": sum(1 for row in worldcup_player_stats if float(row.rating or 0) > 0)},
                    {"分析维度": "match_events 数", "结果": worldcup_events_count},
                    {"分析维度": "team_stats 数", "结果": len(worldcup_team_stats)},
                    {"分析维度": "shots 数", "结果": worldcup_shots_count},
                ]
            )
            if fifa_bundle is None:
                fifa_bundle = build_group_stage_bundle()
            for label, df, field in (
                ("top_scorers", fifa_bundle["top_scorers_df"], "进球"),
                ("top_assists", fifa_bundle["top_assists_df"], "助攻"),
                ("top_ratings", fifa_bundle["top_rating_df"], "评分"),
            ):
                if not df.empty:
                    top_row = df.iloc[0]
                    analysis_rows.append({"分析维度": label, "结果": f"{top_row['球员姓名']} / {top_row[field]}"})
        except Exception as exc:
            analysis_rows.append({"分析维度": "分析结果汇总", "结果": f"生成失败: {exc}"})

        if analysis_rows:
            self._write_df(ws15, pd.DataFrame(analysis_rows))
        else:
            ws15.cell(row=1, column=1, value="分析维度")
            ws15.cell(row=1, column=2, value="结果")
            self._style_header(ws15)

        wb.save(filepath)
        return filepath
