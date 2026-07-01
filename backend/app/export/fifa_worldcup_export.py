"""FIFA World Cup group-stage export helpers backed by HDFS raw snapshots."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.hdfs_client import hdfs_client

PLAYER_INFO_COLUMNS = [
    "阶段",
    "分组",
    "小组排名",
    "球队",
    "球员姓名",
    "位置",
    "号码",
    "国籍",
    "出生日期",
    "年龄",
    "身高(cm)",
    "体重(kg)",
    "球员来源ID",
    "照片URL",
]

PLAYER_STATS_COLUMNS = [
    "阶段",
    "分组",
    "小组排名",
    "球队",
    "球员姓名",
    "位置",
    "号码",
    "出场",
    "进球",
    "助攻",
    "黄牌",
    "红牌",
    "出场时间(分钟)",
    "射门",
    "射正",
    "xG",
    "xA",
    "传球数",
    "传球成功率(%)",
    "抢断",
    "拦截",
    "评分",
    "扑救",
    "扑救成功率(%)",
    "xCS",
    "出击防守",
    "射手榜排名",
    "助攻榜排名",
    "评分排名",
]

STANDINGS_COLUMNS = [
    "阶段",
    "分组",
    "排名",
    "球队",
    "场次",
    "胜",
    "平",
    "负",
    "进球",
    "失球",
    "净胜球",
    "积分",
    "近况",
    "出线状态",
]


def _latest_hdfs_json(hdfs_dir: str) -> tuple[list[dict[str, Any]], str]:
    entries = hdfs_client.client.list(hdfs_dir, status=True)
    if not entries:
        raise FileNotFoundError(f"HDFS directory is empty: {hdfs_dir}")
    latest_name, _ = max(entries, key=lambda item: item[1].get("modificationTime", 0))
    latest_path = f"{hdfs_dir.rstrip('/')}/{latest_name}"
    with hdfs_client.client.read(latest_path) as reader:
        payload = json.loads(reader.read().decode("utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"Expected list payload at {latest_path}")
    return payload, latest_path


def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _compute_age(birth_date: Any, as_of: date | None) -> int | None:
    born = _parse_date(birth_date)
    if not born or not as_of:
        return None
    years = as_of.year - born.year
    if (as_of.month, as_of.day) < (born.month, born.day):
        years -= 1
    return years


def _safe_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _safe_float(value: Any, digits: int = 2) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return None


def _normalize_stage_label() -> str:
    return "世界杯小组赛"


def _build_schedule_metadata(schedule_rows: list[dict[str, Any]]) -> dict[str, Any]:
    group_rows = [row for row in schedule_rows if str(row.get("group") or "").startswith("Group")]
    finished_rows = [row for row in group_rows if row.get("status") == "finished"]
    latest_group_date = None
    parsed_dates = [_parse_date(row.get("date")) for row in finished_rows]
    parsed_dates = [value for value in parsed_dates if value is not None]
    if parsed_dates:
        latest_group_date = max(parsed_dates).isoformat()
    return {
        "group_match_count": len(group_rows),
        "finished_group_match_count": len(finished_rows),
        "group_stage_last_match_date": latest_group_date,
    }


def _build_standings_df(standings_rows: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for row in standings_rows:
        rows.append(
            {
                "阶段": _normalize_stage_label(),
                "分组": row.get("group"),
                "排名": _safe_int(row.get("position")),
                "球队": row.get("team"),
                "场次": _safe_int(row.get("played")),
                "胜": _safe_int(row.get("won")),
                "平": _safe_int(row.get("drawn")),
                "负": _safe_int(row.get("lost")),
                "进球": _safe_int(row.get("goals_for")),
                "失球": _safe_int(row.get("goals_against")),
                "净胜球": _safe_int(row.get("goal_diff")),
                "积分": _safe_int(row.get("points")),
                "近况": row.get("form"),
                "出线状态": row.get("qualification_status"),
            }
        )
    df = pd.DataFrame(rows, columns=STANDINGS_COLUMNS)
    if not df.empty:
        df = df.sort_values(by=["分组", "排名", "球队"], na_position="last").reset_index(drop=True)
    return df


def _build_team_lookup(standings_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for record in standings_df.to_dict(orient="records"):
        team = record.get("球队")
        if team:
            lookup[str(team)] = record
    return lookup


def _build_player_frames(
    player_rows: list[dict[str, Any]],
    team_lookup: dict[str, dict[str, Any]],
    group_stage_last_match_date: str | None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    player_info_rows: list[dict[str, Any]] = []
    player_stats_rows: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    cutoff_date = _parse_date(group_stage_last_match_date)

    for row in player_rows:
        standing = team_lookup.get(str(row.get("team") or ""), {})
        group_name = standing.get("分组")
        group_rank = standing.get("排名")
        common = {
            "阶段": _normalize_stage_label(),
            "分组": group_name,
            "小组排名": group_rank,
            "球队": row.get("team"),
            "球员姓名": row.get("name"),
            "位置": row.get("position"),
            "号码": _safe_int(row.get("shirt_number")),
        }
        player_info_rows.append(
            {
                **common,
                "国籍": row.get("nationality"),
                "出生日期": row.get("birth_date"),
                "年龄": _compute_age(row.get("birth_date"), cutoff_date),
                "身高(cm)": _safe_int(row.get("height")),
                "体重(kg)": _safe_int(row.get("weight")),
                "球员来源ID": row.get("player_source_id"),
                "照片URL": row.get("photo_url"),
            }
        )
        player_stats_rows.append(
            {
                **common,
                "出场": _safe_int(row.get("appearances")) or 0,
                "进球": _safe_int(row.get("goals")) or 0,
                "助攻": _safe_int(row.get("assists")) or 0,
                "黄牌": _safe_int(row.get("yellow_cards")) or 0,
                "红牌": _safe_int(row.get("red_cards")) or 0,
                "出场时间(分钟)": _safe_int(row.get("minutes_played")) or 0,
                "射门": _safe_int(row.get("shots")) or 0,
                "射正": _safe_int(row.get("shots_on_target")) or 0,
                "xG": _safe_float(row.get("xg"), 4) or 0.0,
                "xA": _safe_float(row.get("xa"), 4) or 0.0,
                "传球数": _safe_int(row.get("passes")) or 0,
                "传球成功率(%)": _safe_float(row.get("pass_accuracy"), 2) or 0.0,
                "抢断": _safe_int(row.get("tackles")) or 0,
                "拦截": _safe_int(row.get("interceptions")) or 0,
                "评分": _safe_float(row.get("rating"), 2) or 0.0,
                "扑救": _safe_int(row.get("saves")) or 0,
                "扑救成功率(%)": _safe_float(row.get("save_rate"), 2) or 0.0,
                "xCS": _safe_float(row.get("xcs"), 4) or 0.0,
                "出击防守": _safe_int(row.get("sweeper_actions")) or 0,
            }
        )
        raw_rows.append(
            {
                "group": group_name,
                "group_rank": group_rank,
                **row,
            }
        )

    info_df = pd.DataFrame(player_info_rows, columns=PLAYER_INFO_COLUMNS)
    stats_df = pd.DataFrame(player_stats_rows)
    if not stats_df.empty:
        stats_df = stats_df.sort_values(
            by=["进球", "助攻", "评分", "出场时间(分钟)", "球队", "号码"],
            ascending=[False, False, False, False, True, True],
            na_position="last",
        ).reset_index(drop=True)
        stats_df["射手榜排名"] = stats_df["进球"].rank(method="min", ascending=False).astype("Int64")
        stats_df["助攻榜排名"] = stats_df["助攻"].rank(method="min", ascending=False).astype("Int64")
        stats_df["评分排名"] = stats_df["评分"].rank(method="min", ascending=False).astype("Int64")
        stats_df = stats_df[PLAYER_STATS_COLUMNS]
    raw_df = pd.DataFrame(raw_rows)
    return info_df, stats_df, raw_df


def _build_source_df(metadata: dict[str, Any]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"项目": "导出主题", "值": _normalize_stage_label() + "球员数据附件"},
            {"项目": "导出时间", "值": metadata["exported_at"]},
            {"项目": "球员快照路径", "值": metadata["player_snapshot_path"]},
            {"项目": "积分榜路径", "值": metadata["standings_snapshot_path"]},
            {"项目": "赛程路径", "值": metadata["schedule_snapshot_path"]},
            {"项目": "小组赛赛程场次", "值": metadata["group_match_count"]},
            {"项目": "已结束小组赛场次", "值": metadata["finished_group_match_count"]},
            {"项目": "小组赛最后比赛日", "值": metadata["group_stage_last_match_date"]},
            {"项目": "球员总数", "值": metadata["player_count"]},
            {"项目": "带统计字段球员数", "值": metadata["active_player_count"]},
            {"项目": "说明", "值": "当前导出基于 FIFA 官方 HDFS 原始快照，适用于截至小组赛结束的世界杯球员数据附件。"},
        ]
    )


def _build_top_n_df(stats_df: pd.DataFrame, metric: str, top_n: int = 20) -> pd.DataFrame:
    if stats_df.empty:
        return pd.DataFrame(columns=["排名", "球员姓名", "球队", "位置", metric])
    df = stats_df[["球员姓名", "球队", "位置", metric]].copy()
    df = df.sort_values(by=[metric, "球队", "球员姓名"], ascending=[False, True, True]).head(top_n).reset_index(drop=True)
    df.insert(0, "排名", range(1, len(df) + 1))
    return df


def build_group_stage_bundle() -> dict[str, Any]:
    player_rows, player_snapshot_path = _latest_hdfs_json("/sports/raw/fifa_official/player_stats")
    standings_rows, standings_snapshot_path = _latest_hdfs_json("/sports/raw/fifa_official/standings")
    schedule_rows, schedule_snapshot_path = _latest_hdfs_json("/sports/raw/fifa_official/schedule")

    schedule_meta = _build_schedule_metadata(schedule_rows)
    standings_df = _build_standings_df(standings_rows)
    team_lookup = _build_team_lookup(standings_df)
    player_info_df, player_stats_df, raw_player_df = _build_player_frames(
        player_rows=player_rows,
        team_lookup=team_lookup,
        group_stage_last_match_date=schedule_meta["group_stage_last_match_date"],
    )
    active_player_count = 0
    if not player_stats_df.empty:
        metric_cols = ["出场", "进球", "助攻", "出场时间(分钟)", "传球数", "射门", "扑救"]
        active_player_count = int((player_stats_df[metric_cols].sum(axis=1) > 0).sum())

    metadata = {
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "player_snapshot_path": player_snapshot_path,
        "standings_snapshot_path": standings_snapshot_path,
        "schedule_snapshot_path": schedule_snapshot_path,
        "player_count": len(player_rows),
        "active_player_count": active_player_count,
        **schedule_meta,
    }
    source_df = _build_source_df(metadata)
    top_scorers_df = _build_top_n_df(player_stats_df, "进球")
    top_assists_df = _build_top_n_df(player_stats_df, "助攻")
    top_rating_df = _build_top_n_df(player_stats_df, "评分")

    return {
        "metadata": metadata,
        "source_df": source_df,
        "player_info_df": player_info_df,
        "player_stats_df": player_stats_df,
        "standings_df": standings_df,
        "top_scorers_df": top_scorers_df,
        "top_assists_df": top_assists_df,
        "top_rating_df": top_rating_df,
        "raw_player_df": raw_player_df,
    }


def export_group_stage_csv_bundle(export_dir: str | Path) -> dict[str, Path]:
    bundle = build_group_stage_bundle()
    export_root = Path(export_dir)
    export_root.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    players_csv = export_root / f"worldcup_group_stage_players_{timestamp}.csv"
    stats_csv = export_root / f"worldcup_group_stage_player_stats_{timestamp}.csv"
    standings_csv = export_root / f"worldcup_group_stage_standings_{timestamp}.csv"
    source_csv = export_root / f"worldcup_group_stage_sources_{timestamp}.csv"
    bundle_json = export_root / f"worldcup_group_stage_bundle_{timestamp}.json"

    bundle["player_info_df"].to_csv(players_csv, index=False, encoding="utf-8-sig")
    bundle["player_stats_df"].to_csv(stats_csv, index=False, encoding="utf-8-sig")
    bundle["standings_df"].to_csv(standings_csv, index=False, encoding="utf-8-sig")
    bundle["source_df"].to_csv(source_csv, index=False, encoding="utf-8-sig")

    serializable = {
        "metadata": bundle["metadata"],
        "players": bundle["player_info_df"].to_dict(orient="records"),
        "player_stats": bundle["player_stats_df"].to_dict(orient="records"),
        "standings": bundle["standings_df"].to_dict(orient="records"),
        "sources": bundle["source_df"].to_dict(orient="records"),
    }
    bundle_json.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "players_csv": players_csv,
        "stats_csv": stats_csv,
        "standings_csv": standings_csv,
        "source_csv": source_csv,
        "bundle_json": bundle_json,
    }


def _style_header(ws, row_index: int = 1):
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[row_index]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 28)


def _write_df(ws, df: pd.DataFrame):
    for j, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=col_name)
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    _style_header(ws)
    _auto_width(ws)


def export_group_stage_attachment_workbook(export_dir: str | Path) -> Path:
    bundle = build_group_stage_bundle()
    export_root = Path(export_dir)
    export_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook_path = export_root / f"worldcup_group_stage_attachment_{timestamp}.xlsx"

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "说明"
    _write_df(ws0, bundle["source_df"])

    ws1 = wb.create_sheet("球员信息")
    _write_df(ws1, bundle["player_info_df"])

    ws2 = wb.create_sheet("球员统计")
    _write_df(ws2, bundle["player_stats_df"])

    ws3 = wb.create_sheet("小组积分榜")
    _write_df(ws3, bundle["standings_df"])

    ws4 = wb.create_sheet("射手榜TOP20")
    _write_df(ws4, bundle["top_scorers_df"])

    ws5 = wb.create_sheet("助攻榜TOP20")
    _write_df(ws5, bundle["top_assists_df"])

    ws6 = wb.create_sheet("评分TOP20")
    _write_df(ws6, bundle["top_rating_df"])

    wb.save(workbook_path)
    return workbook_path


def export_group_stage_artifacts(export_dir: str | Path) -> dict[str, Path]:
    outputs = export_group_stage_csv_bundle(export_dir)
    outputs["attachment_xlsx"] = export_group_stage_attachment_workbook(export_dir)
    return outputs
