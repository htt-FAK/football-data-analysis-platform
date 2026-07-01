"""Football-Data 最小可用数据流水线。

职责：
1. 抓取 Football-Data 原始比赛数据
2. 生成清洗前 / 清洗后两份 DataFrame
3. 导出两份 Excel 成品

说明：
- 当前优先服务于项目验证与课程展示，因此先围绕“比赛层”闭环。
- 世界杯等专题数据可在此基础上替换数据源与字段映射，不影响整体导出结构。
"""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.cleaning.dedup import dedup_records
from app.cleaning.field_mapping import map_fields
from app.cleaning.missing_value import DEFAULT_RULES, fill_missing
from app.crawlers.football_data import FootballDataCrawler

CORE_FIELDS = [
    "source", "league", "season", "date", "time", "home_team", "away_team",
    "home_score", "away_score", "result", "home_shots", "away_shots",
    "home_shots_on_target", "away_shots_on_target", "home_yellow", "away_yellow",
    "home_red", "away_red", "home_corners", "away_corners", "home_fouls",
    "away_fouls", "referee",
]

FIELD_LABELS = {
    "source": "数据源", "league": "赛事", "season": "赛季", "date": "比赛日期",
    "time": "比赛时间", "home_team": "主队", "away_team": "客队",
    "home_score": "主队进球", "away_score": "客队进球", "result": "赛果",
    "home_shots": "主队射门", "away_shots": "客队射门",
    "home_shots_on_target": "主队射正", "away_shots_on_target": "客队射正",
    "home_yellow": "主队黄牌", "away_yellow": "客队黄牌",
    "home_red": "主队红牌", "away_red": "客队红牌",
    "home_corners": "主队角球", "away_corners": "客队角球",
    "home_fouls": "主队犯规", "away_fouls": "客队犯规", "referee": "裁判",
}

RESULT_MAP = {"H": "home_win", "D": "draw", "A": "away_win"}
RESULT_LABELS = {"home_win": "主胜", "draw": "平局", "away_win": "客胜"}


def normalize_football_data_record(record: dict, league: str, season: str) -> dict:
    mapped = map_fields(record, "football_data")
    cleaned = deepcopy(mapped)
    cleaned["source"] = "football_data"
    cleaned["league"] = league
    cleaned["season"] = season

    raw_date = cleaned.get("date")
    if raw_date:
        try:
            cleaned["date"] = pd.to_datetime(raw_date, dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            cleaned["date"] = str(raw_date)

    if cleaned.get("result") in RESULT_MAP:
        cleaned["result"] = RESULT_MAP[cleaned["result"]]

    cleaned = fill_missing(cleaned, DEFAULT_RULES)
    return {field: cleaned.get(field) for field in CORE_FIELDS}


def fetch_raw_matches(league_code: str = "E0", season_code: str = "2425") -> list[dict]:
    crawler = FootballDataCrawler()
    return crawler.crawl(target="matches", league=league_code, season=season_code)


def build_clean_matches(raw_records: list[dict], league_name: str = "英超", season_code: str = "2425") -> list[dict]:
    clean_records = [normalize_football_data_record(r, league_name, season_code) for r in raw_records]
    return dedup_records(clean_records, ["date", "time", "home_team", "away_team"])


def _style_header(ws):
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 28)


def _write_df(ws, df: pd.DataFrame):
    for j, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=col_name)
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    _style_header(ws)
    _auto_width(ws)


def _build_summary_df(raw_df: pd.DataFrame, clean_df: pd.DataFrame, league_name: str, season_code: str) -> pd.DataFrame:
    return pd.DataFrame([
        {"项目": "数据源", "值": "football-data.co.uk"},
        {"项目": "赛事", "值": league_name},
        {"项目": "赛季", "值": season_code},
        {"项目": "清洗前记录数", "值": len(raw_df)},
        {"项目": "清洗后记录数", "值": len(clean_df)},
        {"项目": "清洗后字段数", "值": len(clean_df.columns)},
        {"项目": "说明", "值": "清洗后保留比赛层核心字段，便于后端入库与前端展示"},
    ])


def _build_raw_desc_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame([
        {"字段名": col, "说明": "来自 Football-Data 原始 CSV 字段", "示例值": raw_df.iloc[0][col] if len(raw_df) else ""}
        for col in raw_df.columns
    ])


def _build_clean_desc_df() -> pd.DataFrame:
    return pd.DataFrame([
        {"字段名": field, "中文名": FIELD_LABELS.get(field, field), "说明": "清洗后的标准字段"}
        for field in CORE_FIELDS
    ])


def export_raw_and_clean_excel(export_dir: str | Path, league_code: str = "E0", season_code: str = "2425", league_name: str = "英超") -> tuple[Path, Path, int]:
    raw_records = fetch_raw_matches(league_code=league_code, season_code=season_code)
    if not raw_records:
        raise RuntimeError("Football-Data 未抓到任何原始记录")

    export_dir = Path(export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_path = export_dir / f"football_data_raw_{league_code}_{season_code}_{timestamp}.xlsx"
    clean_path = export_dir / f"football_data_clean_{league_code}_{season_code}_{timestamp}.xlsx"

    raw_df = pd.DataFrame(raw_records)
    clean_records = build_clean_matches(raw_records, league_name=league_name, season_code=season_code)
    clean_df = pd.DataFrame(clean_records, columns=CORE_FIELDS)
    clean_df_display = clean_df.rename(columns=FIELD_LABELS)
    if "赛果" in clean_df_display.columns:
        clean_df_display["赛果"] = clean_df_display["赛果"].map(lambda x: RESULT_LABELS.get(x, x))

    summary_df = _build_summary_df(raw_df, clean_df, league_name, season_code)
    raw_desc_df = _build_raw_desc_df(raw_df)
    clean_desc_df = _build_clean_desc_df()

    wb_raw = Workbook()
    ws0 = wb_raw.active
    ws0.title = "说明"
    _write_df(ws0, summary_df)
    ws1 = wb_raw.create_sheet("原始数据")
    _write_df(ws1, raw_df)
    ws2 = wb_raw.create_sheet("原始字段说明")
    _write_df(ws2, raw_desc_df)
    wb_raw.save(raw_path)

    wb_clean = Workbook()
    ws0 = wb_clean.active
    ws0.title = "说明"
    _write_df(ws0, summary_df)
    ws1 = wb_clean.create_sheet("清洗后数据")
    _write_df(ws1, clean_df_display)
    ws2 = wb_clean.create_sheet("清洗后字段说明")
    _write_df(ws2, clean_desc_df)
    wb_clean.save(clean_path)

    return raw_path, clean_path, len(raw_records)
