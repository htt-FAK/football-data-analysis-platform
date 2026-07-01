from __future__ import annotations

import asyncio
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
import sys

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "backend"))

from app.database import Base, get_db
from app.main import app
from app.models.crawl_log import CrawlLog
from app.models.data_source import DataSource
from app.models.league import League
from app.models.match import Match
from app.models.match_event import MatchEvent
from app.models.player import Player
from app.models.player_stat import PlayerStat
from app.models.season import Season
from app.models.shot import Shot
from app.models.standings import Standings
from app.models.team import Team
from app.models.team_stat import TeamStat
from app.scheduler import jobs as scheduler_jobs
from app.api.crawl import _build_crawl_kwargs
from app.api import crawl as crawl_api
from app.crawlers.football_data import FootballDataCrawler
from app.crawlers.openligadb import OpenLigaDBCrawler
from app.crawlers.statsbomb import StatsBombCrawler
from app.crawlers.teamrankings import TeamRankingsCrawler
from app.crawlers.thesportsdb import TheSportsDBCrawler
from app.crawlers.understat import UnderstatCrawler
from app.services.ingest_service import ingest_shots
from app.services.data_source_bootstrap import ensure_builtin_data_sources
from app.services.team_service import TeamService
from app.crawlers.fifa_official import FIFAOfficialCrawler
from app.services import match_service as match_service_module


class FakeRedis:
    def __init__(self, store: dict[str, str] | None = None, fail_ping: bool = False):
        self.store = store or {}
        self.fail_ping = fail_ping

    async def ping(self):
        if self.fail_ping:
            raise RuntimeError("redis unavailable")
        return True

    async def get(self, key: str):
        return self.store.get(key)

    async def set(self, key: str, value: str, ex: int | None = None):
        self.store[key] = value

    async def delete(self, key: str):
        self.store.pop(key, None)

    async def scan_iter(self, match: str = "*", count: int = 100):
        prefix = match.replace("*", "")
        for key in list(self.store.keys()):
            if key.startswith(prefix):
                yield key


engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


@asynccontextmanager
async def noop_lifespan(_app):
    yield


app.dependency_overrides[get_db] = override_get_db
app.router.lifespan_context = noop_lifespan


def seed_demo_data():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    db.add_all(
        [
            League(id=3, name="\u4e16\u754c\u676f", country="World", type="cup"),
            Season(id=4, league_id=3, name="2026"),
            Season(id=5, league_id=3, name="2025"),
            Team(id=10, name="Brazil", country="Brazil"),
            Team(id=11, name="Argentina", country="Argentina"),
            Team(id=12, name="France", country="France"),
            Team(id=13, name="Spain", country="Spain"),
            Player(
                id=100,
                team_id=10,
                name="Player A",
                position="FW",
                overall_rating=88,
                atk_score=90,
                org_score=80,
                def_score=30,
                phy_score=75,
                dis_score=70,
                data_source="fifa_official",
            ),
            Player(
                id=101,
                team_id=11,
                name="Legacy FW",
                position="FW",
                overall_rating=60,
                atk_score=50,
                org_score=45,
                def_score=20,
                phy_score=40,
                dis_score=35,
                data_source="football_data",
            ),
            DataSource(
                id=1,
                source_code="fifa_official",
                name="FIFA Official",
                type="crawler",
                enabled=True,
                status="active",
                priority=1,
            ),
            CrawlLog(
                id=1,
                source_id=1,
                target="statistics",
                status="success",
                fetched=4,
                updated=4,
                failed=0,
            ),
            Match(
                id=1000,
                league_id=3,
                season_id=4,
                home_team_id=10,
                away_team_id=11,
                status="finished",
                match_date=datetime(2026, 6, 20, 20, 0),
                stage="Group Stage",
                group_name="Group A",
                home_score=2,
                away_score=1,
                venue="Demo Stadium",
            ),
            Match(
                id=1001,
                league_id=3,
                season_id=4,
                home_team_id=12,
                away_team_id=13,
                status="scheduled",
                match_date=datetime(2026, 6, 30, 18, 0),
                stage="Round of 32",
                venue="Future Stadium",
            ),
            Match(
                id=1002,
                league_id=3,
                season_id=4,
                home_team_id=None,
                away_team_id=None,
                status="scheduled",
                match_date=datetime(2026, 7, 1, 21, 0),
                stage="Round of 32",
                venue="TBD Stadium",
            ),
            Standings(
                id=1,
                season_id=4,
                team_id=10,
                position=1,
                played=3,
                won=2,
                drawn=1,
                lost=0,
                goals_for=6,
                goals_against=2,
                goal_diff=4,
                points=7,
                group_name="Group A",
                stage="Group Stage",
                qualification_status="Qualified",
            ),
            PlayerStat(
                id=1,
                player_id=100,
                season_id=4,
                appearances=3,
                goals=2,
                assists=1,
                minutes_played=270,
                shots=8,
                shots_on_target=4,
                xg=1.8,
                xa=0.7,
                passes=95,
                pass_accuracy=88.5,
                rating=7.8,
            ),
            PlayerStat(
                id=2,
                player_id=101,
                season_id=5,
                appearances=2,
                goals=1,
                assists=0,
                minutes_played=120,
                shots=3,
                shots_on_target=1,
                xg=0.5,
                passes=20,
                pass_accuracy=70.0,
                rating=6.2,
            ),
            TeamStat(
                id=1,
                team_id=10,
                season_id=4,
                matches_played=3,
                wins=2,
                draws=1,
                losses=0,
                goals_for=6,
                goals_against=2,
                xg_for=5.5,
                xg_against=2.1,
                possession=57.2,
                shots_total=38,
                shots_on_target_total=16,
                passes_total=1200,
                pass_accuracy=87.4,
                corners=12,
                fouls=18,
                clean_sheets=1,
                attack_rating=82.0,
                defense_rating=77.0,
                overall_rating=80.0,
            ),
            MatchEvent(
                id=1,
                match_id=1000,
                minute=12,
                event_type="goal",
                team_id=10,
                player_id=100,
                detail="Open play",
            ),
            MatchEvent(
                id=2,
                match_id=1000,
                minute=55,
                event_type="yellow_card",
                team_id=11,
                detail="Foul",
            ),
            MatchEvent(
                id=3,
                match_id=1000,
                minute=None,
                event_type=None,
                detail=None,
            ),
            Shot(
                id=1,
                match_id=1000,
                team_id=10,
                player_id=100,
                minute=12,
                x_coord=0.8,
                y_coord=0.4,
                result="goal",
                xg=0.35,
                shot_type="right_foot",
                situation="open_play",
            ),
            Shot(
                id=2,
                match_id=1000,
                team_id=11,
                minute=60,
                x_coord=0.7,
                y_coord=0.6,
                result="saved",
                xg=0.12,
                shot_type="left_foot",
                situation="open_play",
            ),
        ]
    )
    db.commit()
    db.close()


def test_worldcup_upcoming_and_coverage():
    seed_demo_data()
    import app.api.worldcup as worldcup_api

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 29, 0, 0, 0, tzinfo=tz)

    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(worldcup_api, "datetime", FrozenDateTime)
    client = TestClient(app)

    try:
        coverage = client.get("/api/v1/worldcup/coverage?season=2026")
        assert coverage.status_code == 200
        coverage_payload = coverage.json()["coverage"]
        key_event_module = next(item for item in coverage_payload if item["module"] == "关键事件影响分析")
        assert key_event_module["status"] == "ready"
        assert key_event_module["detail"] == "match_events=2"

        upcoming = client.get("/api/v1/worldcup/upcoming?season=2026&limit=16")
        assert upcoming.status_code == 200
        payload = upcoming.json()
        assert payload["season"] == "2026"
        assert len(payload["matches"]) == 2
        assert [row["match_id"] for row in payload["matches"]] == [1001, 1002]
        assert payload["matches"][0]["is_ready_for_prediction"] is True
        assert payload["matches"][0]["home_team_name"] == "France"
        assert payload["matches"][1]["is_ready_for_prediction"] is False
        assert payload["matches"][1]["home_team_name"] is None
        assert payload["matches"][1]["away_team_name"] is None
    finally:
        monkeypatch.undo()


def test_worldcup_player_radar_requires_worldcup_player_stats():
    seed_demo_data()
    client = TestClient(app)

    ok_response = client.get("/api/v1/worldcup/players/100/radar?season=2026")
    assert ok_response.status_code == 200
    ok_payload = ok_response.json()
    assert ok_payload["dimensions"] == [
        "Attacking Impact",
        "Chance Creation",
        "Physical Duel",
        "Pressing",
        "Decision Making",
        "Overall Level",
    ]
    assert ok_payload["values"] == [90, 80, 75, 30, 70, 88]

    db = TestingSessionLocal()
    try:
        db.add(
            MatchEvent(
                id=4,
                match_id=1000,
                minute=77,
                event_type="substitution",
                team_id=10,
                player_id=101,
                detail="Legacy player event only",
            )
        )
        db.commit()
    finally:
        db.close()

    missing_stat_response = client.get("/api/v1/worldcup/players/101/radar?season=2026")
    assert missing_stat_response.status_code == 404


def test_player_endpoints_resolve_sparse_event_player_to_worldcup_canonical():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                Team(id=34, name="Brazil", country="Brazil"),
                Player(
                    id=169,
                    team_id=34,
                    name="CASEMIRO",
                    position="MF",
                    nationality="BRA",
                    overall_rating=39.4,
                    atk_score=38.0,
                    org_score=45.0,
                    def_score=67.0,
                    phy_score=73.0,
                    dis_score=62.0,
                    data_source="fifa_official",
                ),
                Player(
                    id=2350,
                    team_id=34,
                    name="卡塞米罗",
                    overall_rating=0.0,
                    data_source="dongqiudi",
                ),
                PlayerStat(
                    id=999,
                    player_id=169,
                    season_id=4,
                    appearances=4,
                    goals=1,
                    assists=0,
                    minutes_played=354,
                    shots=7,
                    shots_on_target=2,
                    xg=0.8,
                    xa=0.1,
                    passes=120,
                    pass_accuracy=89.0,
                    tackles=9,
                    interceptions=5,
                    rating=4.57,
                    data_source="fifa_official",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)

    detail = client.get("/api/v1/players/2350")
    assert detail.status_code == 200
    detail_payload = detail.json()
    assert detail_payload["id"] == 169
    assert detail_payload["requested_id"] == 2350
    assert detail_payload["canonical_id"] == 169
    assert detail_payload["position"] == "MF"
    assert detail_payload["team_name"] == "Brazil"

    stats = client.get("/api/v1/players/2350/stats")
    assert stats.status_code == 200
    stats_payload = stats.json()
    assert stats_payload["player_id"] == 169
    assert stats_payload["requested_player_id"] == 2350
    assert stats_payload["canonical_player_id"] == 169
    assert stats_payload["stats"]["appearances"] == 4

    radar = client.get("/api/v1/players/2350/radar")
    assert radar.status_code == 200
    radar_payload = radar.json()
    assert radar_payload["player_id"] == 169
    assert radar_payload["position"] == "MF"


def test_worldcup_context_prefers_nonempty_alias_league():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                League(id=30, name="FIFA World Cup", country="World", type="cup"),
                Season(id=31, league_id=30, name="2026"),
            ]
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/worldcup/summary?season=2026")
    assert response.status_code == 200
    payload = response.json()
    assert payload["league_id"] == 3
    assert payload["match_count"] == 3
    assert payload["player_count"] == 1


def test_worldcup_upcoming_orders_matches_by_date_then_id():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                Match(
                    id=1003,
                    league_id=3,
                    season_id=4,
                    home_team_id=10,
                    away_team_id=12,
                    status="scheduled",
                    match_date=datetime(2026, 6, 30, 18, 0),
                    stage="Round of 32",
                    venue="Alpha Stadium",
                ),
                Match(
                    id=1004,
                    league_id=3,
                    season_id=4,
                    home_team_id=11,
                    away_team_id=13,
                    status="scheduled",
                    match_date=datetime(2026, 6, 29, 21, 0),
                    stage="Round of 32",
                    venue="Beta Stadium",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    import app.api.worldcup as worldcup_api

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 29, 0, 0, 0, tzinfo=tz)

    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(worldcup_api, "datetime", FrozenDateTime)

    client = TestClient(app)
    try:
        response = client.get("/api/v1/worldcup/upcoming?season=2026&limit=16")
        assert response.status_code == 200

        match_ids = [row["match_id"] for row in response.json()["matches"]]
        assert match_ids == [1004, 1001, 1003, 1002]
    finally:
        monkeypatch.undo()


def test_worldcup_upcoming_excludes_past_scheduled_matches(monkeypatch):
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(
            Match(
                id=1005,
                league_id=3,
                season_id=4,
                home_team_id=10,
                away_team_id=11,
                status="scheduled",
                match_date=datetime(2026, 6, 28, 20, 0),
                stage="Round of 32",
                venue="Expired Stadium",
            )
        )
        db.commit()
    finally:
        db.close()

    import app.api.worldcup as worldcup_api

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 29, 0, 0, 0)

    monkeypatch.setattr(worldcup_api, "datetime", FrozenDateTime)

    client = TestClient(app)
    response = client.get("/api/v1/worldcup/upcoming?season=2026&limit=16")
    assert response.status_code == 200

    match_ids = [row["match_id"] for row in response.json()["matches"]]
    assert 1005 not in match_ids


def test_worldcup_match_endpoints_refresh_stale_fifa_scores(monkeypatch):
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        stale = db.query(Match).filter(Match.id == 1001).first()
        assert stale is not None
        stale.source_id = "400021516"
        stale.data_source = "fifa_official"
        stale.status = "scheduled"
        stale.home_score = None
        stale.away_score = None
        stale.match_date = datetime(2026, 6, 29, 17, 0)
        db.commit()
    finally:
        db.close()

    monkeypatch.setattr(
        match_service_module,
        "_load_worldcup_schedule_index",
        lambda: {
            "400021516": {
                "match_id": "400021516",
                "date": "2026-06-29",
                "time": "17:00",
                "date_time_iso": "2026-06-29T17:00:00Z",
                "league": "FIFA World Cup™",
                "season": "2026",
                "matchday": 76,
                "home_team": "France",
                "home_team_id": "12",
                "away_team": "Spain",
                "away_team_id": "13",
                "home_score": 2,
                "away_score": 1,
                "status": "finished",
                "venue": "Houston Stadium - Houston",
                "stage": "Round of 32",
                "group": None,
                "fdh_match_id": "151681",
                "source": "fifa_official",
            }
        },
    )

    client = TestClient(app)

    match_detail = client.get("/api/v1/matches/1001")
    assert match_detail.status_code == 200
    detail_payload = match_detail.json()
    assert detail_payload["status"] == "finished"
    assert detail_payload["home_score"] == 2
    assert detail_payload["away_score"] == 1

    worldcup_matches = client.get("/api/v1/worldcup/matches?season=2026&limit=16")
    assert worldcup_matches.status_code == 200
    refreshed_match = next(row for row in worldcup_matches.json()["matches"] if row["match_id"] == 1001)
    assert refreshed_match["status"] == "finished"
    assert refreshed_match["home_score"] == 2
    assert refreshed_match["away_score"] == 1

    upcoming = client.get("/api/v1/worldcup/upcoming?season=2026&limit=16")
    assert upcoming.status_code == 200
    upcoming_ids = [row["match_id"] for row in upcoming.json()["matches"]]
    assert 1001 not in upcoming_ids


def test_matches_list_supports_worldcup_group_stage_alias_filters():
    seed_demo_data()
    client = TestClient(app)

    group_stage_response = client.get("/api/v1/matches?league_id=3&stage=Group%20Stage&group=Group%20A&limit=10")
    assert group_stage_response.status_code == 200
    group_stage_payload = group_stage_response.json()
    assert len(group_stage_payload) == 1
    assert group_stage_payload[0]["id"] == 1000

    first_stage_response = client.get("/api/v1/matches?league_id=3&stage=First%20Stage&group=Group%20A&limit=10")
    assert first_stage_response.status_code == 200
    first_stage_payload = first_stage_response.json()
    assert len(first_stage_payload) == 1
    assert first_stage_payload[0]["id"] == 1000

    compat_group_response = client.get("/api/v1/matches?league_id=3&stage=Group%20Stage&group_name=Group%20A&limit=10")
    assert compat_group_response.status_code == 200
    compat_group_payload = compat_group_response.json()
    assert len(compat_group_payload) == 1
    assert compat_group_payload[0]["id"] == 1000


def test_match_and_team_service_endpoints():
    seed_demo_data()
    client = TestClient(app)

    team_stats = client.get("/api/v1/teams/10/stats?season=2026")
    assert team_stats.status_code == 200
    stats_payload = team_stats.json()
    assert stats_payload["stats"]["xg_for"] == 5.5
    assert stats_payload["stats"]["passes_total"] == 1200
    assert stats_payload["xg"] == 5.5
    assert stats_payload["shots"] == 38
    assert stats_payload["overall_score"] == 80.0

    team_shots = client.get("/api/v1/teams/10/shots")
    assert team_shots.status_code == 200
    team_shots_payload = team_shots.json()
    assert isinstance(team_shots_payload, list)
    assert team_shots_payload[0]["x"] == 0.8

    match_events = client.get("/api/v1/matches/1000/events")
    assert match_events.status_code == 200
    event_payload = match_events.json()
    assert isinstance(event_payload, list)
    assert len(event_payload) == 2
    assert [event["minute"] for event in event_payload] == [12, 55]

    match_detail = client.get("/api/v1/matches/1000")
    assert match_detail.status_code == 200
    detail_payload = match_detail.json()
    assert detail_payload["league_name"] == "\u4e16\u754c\u676f"
    assert detail_payload["season"] == "2026"
    assert detail_payload["date_time"] == "2026-06-20T20:00:00"
    assert detail_payload["group_name"] == "Group A"

    match_shots = client.get("/api/v1/matches/1000/shots")
    assert match_shots.status_code == 200
    match_shots_payload = match_shots.json()
    assert isinstance(match_shots_payload, list)
    assert len(match_shots_payload) == 2
    assert match_shots_payload[0]["x"] == 0.8

    xg_timeline = client.get("/api/v1/matches/1000/xg-timeline")
    assert xg_timeline.status_code == 200
    xg_payload = xg_timeline.json()
    assert xg_payload["available"] is True
    assert xg_payload["shot_count"] == 2
    assert xg_payload["home_team"]["final_xg"] == 0.35
    assert xg_payload["away_team"]["final_xg"] == 0.12
    assert xg_payload["note"] is None

    report = client.get("/api/v1/matches/1000/report")
    assert report.status_code == 200
    report_payload = report.json()
    assert len(report_payload["events"]) == 2
    assert report_payload["shots"]["total"] == 2
    assert report_payload["impact_summary"]["key_events_count"] == 2
    assert report_payload["impact_summary"]["key_events"][0]["event_type"] in {"goal", "yellow_card"}
    assert len(report_payload["impact_summary"]["momentum_curve"]) == 2
    assert report_payload["impact_summary"]["momentum_curve"][0]["side"] == "home"


def test_match_xg_timeline_marks_missing_shot_data_as_unavailable():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.query(Shot).delete()
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/matches/1000/xg-timeline")
    assert response.status_code == 200

    payload = response.json()
    assert payload["available"] is False
    assert payload["shot_count"] == 0
    assert payload["home_team"]["final_xg"] is None
    assert payload["away_team"]["final_xg"] is None
    assert payload["timeline"] == {"home": [], "away": []}
    assert "无法生成真实 xG 时间线" in payload["note"]


def test_match_xg_timeline_excludes_incomplete_shot_rows_but_keeps_real_timeline():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(
            Shot(
                id=3,
                match_id=1000,
                team_id=10,
                player_id=100,
                minute=None,
                x_coord=0.5,
                y_coord=0.3,
                result="blocked",
                xg=None,
                shot_type="right_foot",
                situation="set_piece",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/matches/1000/xg-timeline")
    assert response.status_code == 200
    payload = response.json()
    assert payload["available"] is True
    assert payload["shot_count"] == 2
    assert payload["coverage"]["total_rows"] == 3
    assert payload["coverage"]["excluded_rows"] == 1
    assert "未纳入 xG 时间线" in payload["note"]


def test_league_and_data_source_endpoints():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                Team(id=14, name="Sweden", country="Sweden"),
                Team(id=15, name="Mexico", country="Mexico"),
                Match(
                    id=1006,
                    league_id=3,
                    season_id=4,
                    home_team_id=10,
                    away_team_id=15,
                    status="finished",
                    match_date=datetime(2026, 6, 22, 20, 0),
                    stage="Group Stage",
                    group_name="Group A",
                    home_score=1,
                    away_score=1,
                    venue="Trend Stadium 1",
                ),
                Match(
                    id=1007,
                    league_id=3,
                    season_id=4,
                    home_team_id=14,
                    away_team_id=10,
                    status="finished",
                    match_date=datetime(2026, 6, 24, 20, 0),
                    stage="Group Stage",
                    group_name="Group A",
                    home_score=0,
                    away_score=3,
                    venue="Trend Stadium 2",
                ),
                Standings(
                    id=2,
                    season_id=4,
                    team_id=14,
                    position=2,
                    played=3,
                    won=2,
                    drawn=0,
                    lost=1,
                    goals_for=4,
                    goals_against=3,
                    goal_diff=1,
                    points=6,
                    group_name="Group A",
                    stage="Group Stage",
                ),
                Standings(
                    id=3,
                    season_id=4,
                    team_id=15,
                    position=3,
                    played=3,
                    won=1,
                    drawn=1,
                    lost=1,
                    goals_for=3,
                    goals_against=3,
                    goal_diff=0,
                    points=4,
                    group_name="Group A",
                    stage="Group Stage",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)

    league_detail = client.get("/api/v1/leagues/3")
    assert league_detail.status_code == 200
    assert league_detail.json()["name"] == "\u4e16\u754c\u676f"
    assert league_detail.json()["country"] == "World"
    assert league_detail.json()["season"] == "2026"

    world_leagues = client.get("/api/v1/leagues?country=World")
    assert world_leagues.status_code == 200
    assert any(row["id"] == 3 for row in world_leagues.json())

    standings = client.get("/api/v1/leagues/3/standings?season=2026&group=Group%20A")
    assert standings.status_code == 200
    assert standings.json()["standings"][0]["team_name"] == "Brazil"

    schedule = client.get("/api/v1/leagues/3/schedule?season=2026&stage=Group%20Stage")
    assert schedule.status_code == 200
    assert len(schedule.json()["matches"]) == 3

    first_stage_schedule = client.get("/api/v1/leagues/3/schedule?season=2026&stage=First%20Stage")
    assert first_stage_schedule.status_code == 200
    assert len(first_stage_schedule.json()["matches"]) == 3

    first_stage_standings = client.get("/api/v1/leagues/3/standings?season=2026&stage=First%20Stage&group=Group%20A")
    assert first_stage_standings.status_code == 200
    assert len(first_stage_standings.json()["standings"]) == 3

    trends = client.get("/api/v1/leagues/3/trends?season=2026")
    assert trends.status_code == 200
    trends_payload = trends.json()
    assert trends_payload["note"] == "已基于已完赛比赛按时间顺序聚合累计积分趋势"
    assert trends_payload["trends"]
    brazil_trend = next(item for item in trends_payload["trends"] if item["team_name"] == "Brazil")
    assert brazil_trend["current_points"] == 7
    assert brazil_trend["goal_diff"] == 4
    assert len(brazil_trend["timeline"]) == 3
    assert [item["points"] for item in brazil_trend["points_timeline"]] == [3, 4, 7]
    assert brazil_trend["timeline"][0]["result"] == "W"
    assert brazil_trend["timeline"][-1]["opponent_team_name"] == "Sweden"

    health = client.get("/api/v1/data-sources/health")
    assert health.status_code == 200
    assert health.json()[0]["source_code"] == "fifa_official"

    logs = client.get("/api/v1/data-sources/logs")
    assert logs.status_code == 200
    assert logs.json()[0]["target"] == "statistics"

    position_stats = client.get("/api/v1/players/position-stats?position=FW&season=2026")
    assert position_stats.status_code == 200
    position_payload = position_stats.json()
    assert position_payload["count"] == 1

    empty_position_stats = client.get("/api/v1/players/position-stats?position=FW&season=2099")
    assert empty_position_stats.status_code == 200
    empty_position_payload = empty_position_stats.json()
    assert empty_position_payload["count"] == 0
    assert empty_position_payload["distributions"] == {}

    default_position_stats = client.get("/api/v1/players/position-stats?position=FW")
    assert default_position_stats.status_code == 200
    default_position_payload = default_position_stats.json()
    assert default_position_payload["count"] == 1
    assert default_position_payload["distributions"]["atk_score"]["max"] > 0

    position_rank = client.get("/api/v1/players/100/position-rank?season=2026")
    assert position_rank.status_code == 200
    assert position_rank.json()["rank"] == 1

    default_player_stats = client.get("/api/v1/players/100/stats")
    assert default_player_stats.status_code == 200
    assert default_player_stats.json()["season"] == "2026"
    assert default_player_stats.json()["stats"]["goals"] == 2

    default_top_scorers = client.get("/api/v1/players/top-scorers?limit=5")
    assert default_top_scorers.status_code == 200
    assert len(default_top_scorers.json()) == 1
    assert default_top_scorers.json()[0]["player_id"] == 100

    default_compare = client.get("/api/v1/players/compare?player_a=100&player_b=101")
    assert default_compare.status_code == 200
    assert default_compare.json()["season_stats"]["player_a"]["goals"] == 2
    assert default_compare.json()["position_rank"]["player_a"]["rank"] == 1

    missing_season_rank = client.get("/api/v1/players/100/position-rank?season=2099")
    assert missing_season_rank.status_code == 200
    missing_rank_payload = missing_season_rank.json()
    assert missing_rank_payload["rank"] is None
    assert missing_rank_payload["total"] == 0


def test_league_service_group_stage_alias_supports_worldcup_standings():
    seed_demo_data()
    from app.services.league_service import LeagueService

    db = TestingSessionLocal()
    try:
        service = LeagueService()
        group_stage = service.get_standings(db, 3, season="2026", stage="Group Stage", group_name="Group A")
        first_stage = service.get_standings(db, 3, season="2026", stage="First Stage", group_name="Group A")

        assert group_stage is not None
        assert len(group_stage["standings"]) == 1
        assert group_stage["standings"][0]["team_name"] == "Brazil"

        assert first_stage is not None
        assert len(first_stage["standings"]) == 1
        assert first_stage["standings"][0]["team_name"] == "Brazil"
    finally:
        db.close()


def test_data_source_logs_display_empty_success_as_partial():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(
            CrawlLog(
                id=99,
                source_id=1,
                target="players",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/logs?source_id=1&limit=5")
    assert response.status_code == 200
    payload = response.json()
    row = next(item for item in payload if item["id"] == 99)
    assert row["status"] == "partial"
    assert row["raw_status"] == "success"


def test_crawl_status_displays_empty_success_as_partial():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        log = CrawlLog(
            id=98,
            source_id=1,
            target="players",
            start_time=datetime.now(),
            end_time=datetime.now(),
            fetched=0,
            updated=0,
            failed=0,
            status="success",
        )
        db.add(log)
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/crawl/98")
    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "partial"
    assert payload["raw_status"] == "success"


def test_league_list_excludes_empty_shell_leagues():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                League(id=30, name="FIFA World Cup™", country="World", type="cup"),
                Season(id=31, league_id=30, name="2026"),
                League(id=32, name="FIFA World Cup?", country="World", type="cup"),
            ]
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/leagues/")
    assert response.status_code == 200

    rows = response.json()
    league_ids = {row["id"] for row in rows}
    assert 3 in league_ids
    assert 30 not in league_ids
    assert 32 not in league_ids


def test_team_list_respects_requested_season():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(Team(id=14, name="Germany", country="Germany"))
        db.add(
            Match(
                id=1006,
                league_id=3,
                season_id=5,
                home_team_id=14,
                away_team_id=11,
                status="finished",
                match_date=datetime(2025, 6, 18, 20, 0),
                stage="Group Stage",
                group_name="Group B",
                home_score=1,
                away_score=0,
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/teams?league_id=3&season=2026")
    assert response.status_code == 200

    team_names = {row["name"] for row in response.json()}
    assert "Germany" not in team_names
    assert {"Brazil", "Argentina", "France", "Spain"}.issubset(team_names)

    missing_season_response = client.get("/api/v1/teams?league_id=3&season=2099")
    assert missing_season_response.status_code == 200
    assert missing_season_response.json() == []


def test_team_shots_respect_requested_season():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(
            Match(
                id=1007,
                league_id=3,
                season_id=5,
                home_team_id=10,
                away_team_id=11,
                status="finished",
                match_date=datetime(2025, 6, 19, 20, 0),
                stage="Group Stage",
                group_name="Group A",
                home_score=1,
                away_score=0,
            )
        )
        db.add(
            Shot(
                id=3,
                match_id=1007,
                team_id=10,
                player_id=100,
                minute=44,
                x_coord=0.25,
                y_coord=0.75,
                result="blocked",
                xg=0.08,
                shot_type="left_foot",
                situation="open_play",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)

    response_2026 = client.get("/api/v1/teams/10/shots?season=2026")
    assert response_2026.status_code == 200
    payload_2026 = response_2026.json()
    assert len(payload_2026) == 1
    assert payload_2026[0]["match_id"] == 1000

    response_2025 = client.get("/api/v1/teams/10/shots?season=2025")
    assert response_2025.status_code == 200
    payload_2025 = response_2025.json()
    assert len(payload_2025) == 1
    assert payload_2025[0]["match_id"] == 1007


def test_player_compare_key_events_respect_requested_season():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(
            Match(
                id=1008,
                league_id=3,
                season_id=5,
                home_team_id=10,
                away_team_id=11,
                status="finished",
                match_date=datetime(2025, 6, 21, 20, 0),
                stage="Group Stage",
                group_name="Group A",
                home_score=2,
                away_score=0,
            )
        )
        db.add(
            MatchEvent(
                id=4,
                match_id=1008,
                minute=88,
                event_type="red_card",
                team_id=10,
                player_id=100,
                detail="Legacy season red card",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/players/compare?player_a=100&player_b=101&season=2026")
    assert response.status_code == 200
    payload = response.json()

    player_a_events = payload["key_events"]["player_a"]
    assert len(player_a_events) == 1
    assert player_a_events[0]["match_id"] == 1000
    assert player_a_events[0]["event_type"] == "goal"


def test_player_radar_and_compare_degrade_cleanly_for_missing_season():
    seed_demo_data()
    client = TestClient(app)

    radar_response = client.get("/api/v1/players/100/radar?season=2099")
    assert radar_response.status_code == 200
    radar_payload = radar_response.json()
    assert radar_payload["season"] is None
    assert radar_payload["mode"] == "summary_only"
    assert radar_payload["overall"] == 0.0
    assert radar_payload["values"] == [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
    assert radar_payload["completeness"]["label"] == "no_stats"

    compare_response = client.get("/api/v1/players/compare?player_a=100&player_b=101&season=2099")
    assert compare_response.status_code == 200
    compare_payload = compare_response.json()
    assert compare_payload["season_stats"]["player_a"] is None
    assert compare_payload["season_stats"]["player_b"] is None
    assert compare_payload["radar"]["player_a"]["mode"] == "summary_only"
    assert compare_payload["radar"]["player_b"]["mode"] == "summary_only"
    assert compare_payload["radar"]["player_a"]["overall"] == 0.0
    assert compare_payload["radar"]["player_b"]["overall"] == 0.0
    assert compare_payload["recommended_visualization"] == "summary_only"


def test_crawl_trigger_normalizes_statistics_target(monkeypatch):
    seed_demo_data()
    client = TestClient(app)

    calls: list[dict] = []

    def fake_execute(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr("app.api.crawl._execute_crawl", fake_execute)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "fifa_official", "target": "team_stats"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["target"] == "statistics"

    db = TestingSessionLocal()
    try:
        logs = db.query(CrawlLog).order_by(CrawlLog.id.asc()).all()
        assert logs[-1].target == "statistics"
    finally:
        db.close()

    assert calls
    assert calls[0]["target"] == "statistics"


def test_crawl_trigger_supports_shots_target(monkeypatch):
    seed_demo_data()
    client = TestClient(app)

    calls: list[dict] = []

    def fake_execute(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr("app.api.crawl._execute_crawl", fake_execute)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "statsbomb", "target": "shots", "league_name": "FIFA World Cup", "season": "2022"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["target"] == "shots"
    assert calls
    assert calls[0]["target"] == "shots"


def test_crawl_trigger_rejects_unsupported_source_target_combination():
    seed_demo_data()
    client = TestClient(app)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "thesportsdb", "target": "standings"},
    )
    assert response.status_code == 400
    assert "不支持目标" in response.json()["detail"]


def test_crawl_trigger_passes_statsbomb_worldcup_context_for_shots(monkeypatch):
    seed_demo_data()
    client = TestClient(app)

    calls: list[dict] = []

    def fake_execute(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr("app.api.crawl._execute_crawl", fake_execute)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "statsbomb", "target": "shots", "league_name": "世界杯", "season": "2026"},
    )
    assert response.status_code == 400
    assert "does not provide" in response.json()["detail"]


def test_crawl_trigger_rejects_statsbomb_shots_without_explicit_context():
    seed_demo_data()
    client = TestClient(app)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "statsbomb", "target": "shots"},
    )
    assert response.status_code == 400
    assert "requires explicit league and season context" in response.json()["detail"]


def test_crawl_trigger_rejects_unsupported_understat_worldcup_shots():
    seed_demo_data()
    client = TestClient(app)

    response = client.post(
        "/api/v1/crawl/trigger",
        json={"source": "understat", "target": "shots", "league_name": "World Cup", "season": "2026"},
    )
    assert response.status_code == 400
    assert "does not support league" in response.json()["detail"]


def test_build_crawl_kwargs_routes_source_specific_league_params():
    api_kwargs = _build_crawl_kwargs("api_football", "fixtures", "39", "2026", None)
    assert api_kwargs["league_id"] == 39
    assert "league" not in api_kwargs

    sportsdb_kwargs = _build_crawl_kwargs("thesportsdb", "events", "4328", "2025-2026", None)
    assert sportsdb_kwargs["league_id"] == "4328"
    assert "league" not in sportsdb_kwargs

    fifa_kwargs = _build_crawl_kwargs("fifa_official", "schedule", None, "2026", "世界杯")
    assert fifa_kwargs["league"] == "世界杯"
    assert "league_id" not in fifa_kwargs

    statsbomb_kwargs = _build_crawl_kwargs("statsbomb", "shots", None, "2026", "世界杯")
    assert statsbomb_kwargs["league"] == "世界杯"
    assert statsbomb_kwargs["season"] == "2026"


def test_scheduler_registers_worldcup_refresh_job(monkeypatch):
    original_scheduler = scheduler_jobs.scheduler
    test_scheduler = scheduler_jobs.AsyncIOScheduler()
    monkeypatch.setattr(scheduler_jobs, "scheduler", test_scheduler)

    scheduler_jobs.setup_jobs()

    jobs_by_id = {job.id: job for job in test_scheduler.get_jobs()}
    assert "worldcup_schedule_refresh" in jobs_by_id
    assert jobs_by_id["worldcup_schedule_refresh"].trigger.interval.total_seconds() == 900
    assert "daily_crawl" in jobs_by_id
    assert "live_crawl" in jobs_by_id

    if test_scheduler.running:
        test_scheduler.shutdown(wait=False)
    monkeypatch.setattr(scheduler_jobs, "scheduler", original_scheduler)


def test_refresh_worldcup_schedule_only_runs_fifa_schedule(monkeypatch):
    seed_demo_data()

    captured: list[tuple[str, str, str]] = []

    async def fake_crawl_source_target(source, db, target, crawl_target, ingest_fn):
        captured.append((source.source_code, target, crawl_target))

    monkeypatch.setattr(scheduler_jobs, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(scheduler_jobs, "ensure_builtin_data_sources", lambda db: 0)
    monkeypatch.setattr(scheduler_jobs, "_crawl_source_target", fake_crawl_source_target)

    asyncio.run(scheduler_jobs.refresh_worldcup_schedule())

    assert captured == [("fifa_official", "worldcup_schedule", "schedule")]


def test_source_strategy_supports_only_declared_tasks():
    from app.services.source_strategy import resolve_crawl_target, supports_task

    assert supports_task("fifa_official", "match_catalog") is True
    assert supports_task("fifa_official", "player_basic") is True
    assert supports_task("football_data", "match_catalog") is True
    assert supports_task("football_data", "player_basic") is False
    assert supports_task("fbref", "match_catalog") is False
    assert supports_task("fbref", "player_basic") is False
    assert supports_task("fbref", "player_advanced") is True
    assert supports_task("openligadb", "live_match") is False
    assert supports_task("thesportsdb", "player_basic") is False
    assert resolve_crawl_target("football_data", "schedule") == "matches"
    assert resolve_crawl_target("api_football", "schedule") == "fixtures"
    assert resolve_crawl_target("thesportsdb", "players") == "players"
    assert resolve_crawl_target("thesportsdb", "standings") is None


def test_daily_full_crawl_filters_sources_by_supported_task(monkeypatch):
    seed_demo_data()

    calls: list[tuple[str, str, str]] = []

    async def fake_crawl_source_target(source, db, target, crawl_target, ingest_fn):
        calls.append((source.source_code, target, crawl_target))

    monkeypatch.setattr(scheduler_jobs, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(scheduler_jobs, "ensure_builtin_data_sources", lambda db: 0)
    monkeypatch.setattr(scheduler_jobs, "_crawl_source_target", fake_crawl_source_target)

    class DummyRatingService:
        def refresh(self, db, season_name):
            return {"updated": 0}

    monkeypatch.setattr("app.services.worldcup_player_rating_service.WorldCupPlayerRatingService", DummyRatingService)

    asyncio.run(scheduler_jobs.daily_full_crawl())

    schedule_sources = {source for source, target, _ in calls if target == "schedule"}
    standings_sources = {source for source, target, _ in calls if target == "standings"}
    players_sources = {source for source, target, _ in calls if target == "players"}

    assert schedule_sources == {"fifa_official", "api_football", "dongqiudi", "football_data", "openligadb"} or schedule_sources == {"fifa_official"}
    assert "fbref" not in schedule_sources
    assert "football_data" not in players_sources
    assert "fbref" not in players_sources
    assert "thesportsdb" not in players_sources
    assert "fifa_official" in players_sources
    assert "fifa_official" in standings_sources


def test_scheduler_target_mapping_uses_concrete_crawler_target(monkeypatch):
    seed_demo_data()

    captured: list[str] = []

    async def fake_dispatch(source, target):
        captured.append(target)
        return []

    monkeypatch.setattr(scheduler_jobs, "_dispatch_crawl", fake_dispatch)
    monkeypatch.setattr(scheduler_jobs, "SessionLocal", TestingSessionLocal)

    db = TestingSessionLocal()
    try:
        ensure_builtin_data_sources(db)
    finally:
        db.close()

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "football_data").first()
        assert source is not None
        asyncio.run(
            scheduler_jobs._crawl_source_target(
                source,
                db,
                "schedule",
                "schedule",
                lambda *args, **kwargs: {"created": 0, "updated": 0, "failed": 0},
            )
        )
    finally:
        db.close()

    assert captured == ["matches"]


def test_execute_crawl_updates_data_source_status(monkeypatch):
    seed_demo_data()

    class DummyCrawler:
        def crawl(self, **kwargs):
            return [{"team": "Brazil", "played": 3, "won": 2, "drawn": 1, "lost": 0, "goals_for": 6, "goals_against": 2}]

    monkeypatch.setattr(crawl_api, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(crawl_api, "_instantiate_crawler", lambda source: DummyCrawler())
    monkeypatch.setattr(crawl_api, "ingest_team_stats", lambda *args, **kwargs: {"created": 1, "updated": 0, "failed": 0})

    db = TestingSessionLocal()
    try:
        log = CrawlLog(source_id=1, target="statistics", start_time=datetime.now(), status="running")
        db.add(log)
        db.commit()
        db.refresh(log)
        log_id = log.id
    finally:
        db.close()

    crawl_api._execute_crawl(
        log_id=log_id,
        source="fifa_official",
        target="statistics",
        league_code=None,
        season="2026",
        league_name="世界杯",
        season_name="2026",
    )

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        log = db.query(CrawlLog).filter(CrawlLog.id == log_id).first()
        assert source is not None
        assert source.status == "active"
        assert source.last_crawl_at is not None
        assert source.error_count == 0
        assert log is not None
        assert log.status == "success"
    finally:
        db.close()


def test_execute_crawl_marks_empty_success_like_runs_as_partial(monkeypatch):
    seed_demo_data()

    class DummyCrawler:
        def crawl(self, **kwargs):
            return []

    monkeypatch.setattr(crawl_api, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(crawl_api, "_instantiate_crawler", lambda source: DummyCrawler())

    db = TestingSessionLocal()
    try:
        log = CrawlLog(source_id=1, target="statistics", start_time=datetime.now(), status="running")
        db.add(log)
        db.commit()
        db.refresh(log)
        log_id = log.id
    finally:
        db.close()

    crawl_api._execute_crawl(
        log_id=log_id,
        source="fifa_official",
        target="statistics",
        league_code=None,
        season="2026",
        league_name="世界杯",
        season_name="2026",
    )

    db = TestingSessionLocal()
    try:
        log = db.query(CrawlLog).filter(CrawlLog.id == log_id).first()
        assert log is not None
        assert log.fetched == 0
        assert log.updated == 0
        assert log.failed == 0
        assert log.status == "partial"
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        assert source.status == "warning"
    finally:
        db.close()


def test_execute_crawl_marks_source_warning_on_failure(monkeypatch):
    seed_demo_data()

    class DummyCrawler:
        def crawl(self, **kwargs):
            raise RuntimeError("manual crawl failed")

    monkeypatch.setattr(crawl_api, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(crawl_api, "_instantiate_crawler", lambda source: DummyCrawler())

    db = TestingSessionLocal()
    try:
        log = CrawlLog(source_id=1, target="statistics", start_time=datetime.now(), status="running")
        db.add(log)
        db.commit()
        db.refresh(log)
        log_id = log.id
    finally:
        db.close()

    crawl_api._execute_crawl(
        log_id=log_id,
        source="fifa_official",
        target="statistics",
        league_code=None,
        season="2026",
        league_name="世界杯",
        season_name="2026",
    )

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        log = db.query(CrawlLog).filter(CrawlLog.id == log_id).first()
        assert source is not None
        assert source.status == "warning"
        assert source.error_count == 1
        assert log is not None
        assert log.status == "failed"
        assert "manual crawl failed" in (log.error_msg or "")
    finally:
        db.close()


def test_data_source_health_treats_empty_partial_runs_as_warning():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.error_count = 0
        db.add(
            CrawlLog(
                source_id=1,
                target="schedule",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=0,
                updated=0,
                failed=0,
                status="partial",
            )
        )
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        fifa = next(item for item in payload if item["source_code"] == "fifa_official")
        assert fifa["health"] == "warning"
        assert fifa["last_log"]["status"] == "partial"
        assert fifa["last_log"]["fetched"] == 0
    finally:
        db.close()


def test_data_source_health_marks_idle_sources_warning_when_latest_log_is_partial():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        source.last_crawl_at = None
        db.add(
            CrawlLog(
                source_id=1,
                target="schedule",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=0,
                updated=0,
                failed=0,
                status="partial",
            )
        )
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        source_row = next(item for item in payload if item["source_code"] == "fifa_official")
        assert source_row["health"] == "warning"
        assert source_row["last_log"]["status"] == "partial"
    finally:
        db.close()


def test_data_source_health_treats_nonempty_partial_runs_as_warning():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.error_count = 0
        db.add(
            CrawlLog(
                source_id=1,
                target="players",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=10,
                updated=8,
                failed=2,
                status="partial",
            )
        )
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        fifa = next(item for item in payload if item["source_code"] == "fifa_official")
        assert fifa["health"] == "warning"
        assert fifa["last_log"]["status"] == "partial"
        assert fifa["last_log"]["failed"] == 2
    finally:
        db.close()


def test_data_source_health_treats_legacy_empty_success_runs_as_warning():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.error_count = 0
        db.add(
            CrawlLog(
                source_id=1,
                target="players",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        fifa = next(item for item in payload if item["source_code"] == "fifa_official")
        assert fifa["status"] == "warning"
        assert fifa["health"] == "warning"
        assert fifa["last_log"]["status"] == "partial"
        assert fifa["last_log"]["raw_status"] == "success"
        assert fifa["last_log"]["updated"] == 0
    finally:
        db.close()


def test_data_source_health_respects_warning_status_without_logs():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.query(CrawlLog).delete()
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.status = "warning"
        source.error_count = 0
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        fifa = next(item for item in payload if item["source_code"] == "fifa_official")
        assert fifa["health"] == "warning"
        assert fifa["last_log"] is None
    finally:
        db.close()


def test_data_source_health_ignores_legacy_empty_logs_after_source_reset():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        source.last_crawl_at = None
        db.add(
            CrawlLog(
                source_id=1,
                target="players",
                start_time=datetime.now(),
                end_time=datetime.now(),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()

        from app.services.data_source_service import DataSourceService

        payload = DataSourceService().get_health_status(db)
        fifa = next(item for item in payload if item["source_code"] == "fifa_official")
        assert fifa["health"] == "idle"
        assert fifa["last_log"]["status"] == "partial"
        assert fifa["last_log"]["raw_status"] == "success"
    finally:
        db.close()


def test_scheduler_partial_runs_set_source_warning(monkeypatch):
    seed_demo_data()

    async def fake_dispatch(source, target):
        assert target == "schedule"
        return []

    monkeypatch.setattr(scheduler_jobs, "_dispatch_crawl", fake_dispatch)
    monkeypatch.setattr(scheduler_jobs, "SessionLocal", TestingSessionLocal)

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
    finally:
        db.close()

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        asyncio.run(scheduler_jobs._crawl_source_target(source, db, "schedule", "schedule", lambda *args, **kwargs: {"created": 0, "updated": 0, "failed": 0}))
    finally:
        db.close()

    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        log = db.query(CrawlLog).order_by(CrawlLog.id.desc()).first()
        assert source is not None
        assert source.status == "warning"
        assert log is not None
        assert log.status == "partial"
    finally:
        db.close()


def test_scheduler_mark_source_error_sets_warning_before_error_threshold():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.id == 1).first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        db.commit()

        scheduler_jobs._mark_source_error(source, db)

        refreshed = db.query(DataSource).filter(DataSource.id == 1).first()
        assert refreshed is not None
        assert refreshed.status == "warning"
        assert refreshed.error_count == 1
    finally:
        db.close()


def test_scheduler_dispatch_supports_statsbomb_shots(monkeypatch):
    class DummyCrawler:
        def crawl(self, target):
            assert target == "shots"
            return [{"shot_id": "sb-1"}]

    monkeypatch.setitem(sys.modules, "app.crawlers.statsbomb", type("StatsBombModule", (), {"StatsBombCrawler": DummyCrawler}))

    class DummySource:
        source_code = "statsbomb"

    result = asyncio.run(scheduler_jobs._dispatch_crawl(DummySource(), "shots"))
    assert result == [{"shot_id": "sb-1"}]


def test_scheduler_dispatch_propagates_crawler_failures(monkeypatch):
    class DummyCrawler:
        def crawl(self, target):
            raise RuntimeError(f"boom:{target}")

    monkeypatch.setitem(sys.modules, "app.crawlers.statsbomb", type("StatsBombModule", (), {"StatsBombCrawler": DummyCrawler}))

    class DummySource:
        source_code = "statsbomb"

    try:
        asyncio.run(scheduler_jobs._dispatch_crawl(DummySource(), "shots"))
    except RuntimeError as exc:
        assert str(exc) == "boom:shots"
    else:
        raise AssertionError("expected scheduler dispatch to propagate crawler failure")


def test_statsbomb_rejects_unsupported_season():
    crawler = StatsBombCrawler()
    try:
        crawler.crawl(target="shots", league="FIFA World Cup", season="2026")
    except ValueError as exc:
        assert "does not provide" in str(exc)
    else:
        raise AssertionError("expected unsupported StatsBomb season to raise ValueError")


def test_statsbomb_rejects_unsupported_worldcup_alias_season():
    crawler = StatsBombCrawler()
    try:
        crawler.crawl(target="shots", league="世界杯", season="2026")
    except ValueError as exc:
        assert "does not provide" in str(exc)
    else:
        raise AssertionError("expected unsupported StatsBomb world cup alias season to raise ValueError")


def test_understat_rejects_unsupported_worldcup_shots():
    crawler = UnderstatCrawler()
    try:
        crawler.crawl(target="shots", league="World Cup", season="2026")
    except ValueError as exc:
        assert "does not support league" in str(exc)
    else:
        raise AssertionError("expected unsupported Understat league to raise ValueError")


def test_statsbomb_shots_use_match_date_not_event_clock(monkeypatch):
    crawler = StatsBombCrawler()

    def fake_fetch_json(path, timeout=30):
        if path == "matches/43/106.json":
            return [
                {
                    "match_id": 1,
                    "match_date": "2022-12-01",
                }
            ]
        if path == "events/1.json":
            return [
                {
                    "id": "evt-1",
                    "period": 1,
                    "minute": 2,
                    "second": 5,
                    "timestamp": "00:02:05.857",
                    "type": {"name": "Shot"},
                    "team": {"name": "Canada"},
                    "player": {"id": 10, "name": "Shooter"},
                    "shot": {
                        "statsbomb_xg": 0.12,
                        "outcome": {"name": "Saved"},
                        "body_part": {"name": "Right Foot"},
                        "type": {"name": "Open Play"},
                    },
                    "location": [90.0, 30.0],
                }
            ]
        return None

    monkeypatch.setattr(crawler, "_fetch_json", fake_fetch_json)

    rows = crawler.crawl(target="shots", league="43:106", season="2022")
    assert len(rows) == 1
    assert rows[0]["date"] == "2022-12-01"
    assert rows[0]["event_timestamp"] == "00:02:05.857"


def test_ingest_shots_matches_existing_fixture_by_context():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        stats = ingest_shots(
            db,
            [
                {
                    "source": "statsbomb",
                    "match_id_sb": 999001,
                    "home_team": "Brazil",
                    "away_team": "Argentina",
                    "date": "2026-06-20T20:00:00",
                    "team": "Brazil",
                    "player_name": "Player A",
                    "minute": 22,
                    "x_coord_100": 88.0,
                    "y_coord_100": 42.0,
                    "result": "Goal",
                    "shot_type": "Right Foot",
                    "situation": "Open Play",
                    "xg": 0.41,
                    "source_id": "sb-shot-1",
                }
            ],
            source="statsbomb",
        )
        assert stats["created"] == 1

        inserted = db.query(Shot).filter(Shot.source_id == "sb-shot-1").first()
        assert inserted is not None
        assert inserted.match_id == 1000
        assert inserted.team_id == 10
        assert inserted.player_id == 100
    finally:
        db.close()


def test_live_endpoints_and_degradation(monkeypatch):
    seed_demo_data()
    fake_redis = FakeRedis(store={"live:1000": '{"match_id": 1000, "status": "playing"}'})
    monkeypatch.setattr("app.services.live_service.get_redis", lambda: fake_redis)

    client = TestClient(app)

    live_list = client.get("/api/v1/live/")
    assert live_list.status_code == 200
    assert live_list.json()[0]["match_id"] == 1000

    live_one = client.get("/api/v1/live/1000")
    assert live_one.status_code == 200
    assert live_one.json()["status"] == "playing"

    status = client.get("/api/v1/live/status")
    assert status.status_code == 200
    assert status.json()["degraded"] is False
    assert status.json()["redis_available"] is True
    assert status.json()["cache_state"] == "ready"
    assert status.json()["live_match_count"] == 1
    assert status.json()["active_match_ids"] == [1000]

    degraded_redis = FakeRedis(fail_ping=True)
    monkeypatch.setattr("app.services.live_service.get_redis", lambda: degraded_redis)
    degraded = client.get("/api/v1/live/status")
    assert degraded.status_code == 200
    assert degraded.json()["degraded"] is True
    assert degraded.json()["reason"] == "redis_unavailable"
    assert degraded.json()["redis_available"] is False
    assert degraded.json()["cache_state"] == "unavailable"

    degraded_list = client.get("/api/v1/live/")
    assert degraded_list.status_code == 200
    assert degraded_list.json() == []

    degraded_one = client.get("/api/v1/live/1000")
    assert degraded_one.status_code == 200
    assert degraded_one.json() == {}


def test_live_status_distinguishes_empty_cache_from_redis_failure(monkeypatch):
    seed_demo_data()
    fake_redis = FakeRedis()
    monkeypatch.setattr("app.services.live_service.get_redis", lambda: fake_redis)
    client = TestClient(app)

    response = client.get("/api/v1/live/status")
    assert response.status_code == 200
    payload = response.json()
    assert payload["degraded"] is False
    assert payload["reason"] == "no_live_matches"
    assert payload["redis_available"] is True
    assert payload["environment_ready"] is True
    assert payload["cache_state"] == "empty"
    assert payload["live_match_count"] == 0


def test_websocket_subscribe_ack_and_timeout_cleanup(monkeypatch):
    seed_demo_data()
    import app.api.websocket as websocket_api

    original_interval = websocket_api.WS_HEARTBEAT_INTERVAL
    monkeypatch.setattr(websocket_api, "WS_HEARTBEAT_INTERVAL", 1)

    client = TestClient(app)
    try:
        with client.websocket_connect("/ws") as websocket:
            websocket.send_json({"action": "subscribe", "league_ids": [3]})
            payload = websocket.receive_json()
            assert payload["type"] == "ack"
            assert payload["action"] == "subscribe"
            assert payload["league_ids"] == [3]
            assert len(websocket_api.manager.conn_leagues) == 1
            assert 3 in websocket_api.manager.league_subs

            try:
                websocket.receive_json()
            except Exception:
                pass

        assert len(websocket_api.manager.conn_leagues) == 0
        assert 3 not in websocket_api.manager.league_subs
    finally:
        monkeypatch.setattr(websocket_api, "WS_HEARTBEAT_INTERVAL", original_interval)


def test_push_live_update_keeps_all_ongoing_statuses_in_cache(monkeypatch):
    from app.services.ingest_service import push_live_update

    for live_status in ("playing", "live", "in_progress", "half_time"):
        fake_redis = FakeRedis()
        monkeypatch.setattr("app.services.live_service.get_redis", lambda redis=fake_redis: redis)

        asyncio.run(
            push_live_update(
                match_id=1000,
                league_id=3,
                match_data={"match_id": 1000, "status": live_status},
            )
        )

        assert "live:1000" in fake_redis.store

    fake_redis = FakeRedis(store={"live:1000": '{"match_id": 1000, "status": "playing"}'})
    monkeypatch.setattr("app.services.live_service.get_redis", lambda: fake_redis)

    asyncio.run(
        push_live_update(
            match_id=1000,
            league_id=3,
            match_data={"match_id": 1000, "status": "finished"},
        )
    )

    assert "live:1000" not in fake_redis.store


def test_push_live_update_keeps_websocket_flow_alive_when_redis_fails(monkeypatch):
    from app.services.ingest_service import push_live_update

    fake_redis = FakeRedis(fail_ping=True)
    monkeypatch.setattr("app.services.live_service.get_redis", lambda: fake_redis)

    published = []

    class DummyManager:
        async def publish_to_league(self, league_id, message):
            published.append((league_id, message))

    monkeypatch.setattr("app.api.websocket.manager", DummyManager())

    asyncio.run(
        push_live_update(
            match_id=1000,
            league_id=3,
            match_data={"match_id": 1000, "status": "live"},
        )
    )

    assert published
    assert published[0][0] == 3
    assert published[0][1]["data"]["match_id"] == 1000


def test_football_data_crawler_normalizes_match_rows():
    crawler = FootballDataCrawler()
    row = crawler._normalize_match_record(
        {
            "Date": "16/08/2025",
            "Time": "20:00",
            "HomeTeam": "Liverpool",
            "AwayTeam": "Arsenal",
            "FTHG": 2,
            "FTAG": 1,
            "FTR": "H",
            "HS": 14,
            "AS": 9,
            "HST": 6,
            "AST": 3,
            "Referee": "A Ref",
        },
        "E0",
        "2526",
    )
    assert row["league"] == "EPL"
    assert row["date"] == "2025-08-16"
    assert row["home_team"] == "Liverpool"
    assert row["away_team"] == "Arsenal"
    assert row["home_score"] == 2
    assert row["status"] == "finished"


def test_thesportsdb_crawler_normalizes_events_teams_and_players():
    crawler = TheSportsDBCrawler()
    event = crawler._normalize_event(
        {
            "idEvent": "200",
            "idLeague": "4328",
            "strLeague": "English Premier League",
            "strSeason": "2025-2026",
            "strHomeTeam": "Chelsea",
            "strAwayTeam": "Tottenham",
            "intHomeScore": "2",
            "intAwayScore": "2",
            "intRound": "3",
            "strTimestamp": "2025-08-20T19:00:00+00:00",
            "strStatus": "Match Finished",
            "strVenue": "Bridge",
        },
        "4328",
        "2025-2026",
    )
    team = crawler._normalize_team(
        {
            "idTeam": "10",
            "idLeague": "4328",
            "strLeague": "English Premier League",
            "strTeam": "Chelsea",
            "strTeamShort": "CHE",
            "strCountry": "England",
            "strStadium": "Bridge",
            "strManager": "Coach",
            "intFormedYear": "1905",
            "strBadge": "https://badge",
        },
        "4328",
    )
    player = crawler._normalize_player(
        {
            "idPlayer": "99",
            "idTeam": "10",
            "strTeam": "Chelsea",
            "strPlayer": "Demo Player",
            "strPosition": "Forward",
            "strNationality": "England",
            "dateBorn": "2001-01-01",
            "strHeight": "180 cm",
            "strWeight": "75 kg",
            "strThumb": "https://thumb",
        },
        "10",
    )
    assert event["status"] == "finished"
    assert event["home_score"] == 2
    assert event["date"] == "2025-08-20"
    assert team["founded_year"] == 1905
    assert team["name"] == "Chelsea"
    assert player["height"] == 180
    assert player["birth_date"] == "2001-01-01"


def test_openligadb_crawler_normalizes_matches_and_teams():
    crawler = OpenLigaDBCrawler()
    match = crawler._normalize_match(
        {
            "matchID": 55,
            "matchDateTime": "2025-08-23T15:30:00",
            "leagueName": "Bundesliga",
            "leagueShortcut": "bl1",
            "leagueSeason": 2025,
            "group": {"groupOrderID": 2, "groupName": "2. Spieltag"},
            "team1": {"teamId": 1, "teamName": "Bayern"},
            "team2": {"teamId": 2, "teamName": "Dortmund"},
            "matchIsFinished": True,
            "matchResults": [{"resultTypeID": 2, "pointsTeam1": 3, "pointsTeam2": 1}],
            "location": {"locationStadium": "Arena"},
        },
        "bl1",
        "2025",
    )
    team = crawler._normalize_team(
        {
            "teamId": 1,
            "teamName": "Bayern",
            "shortName": "FCB",
            "teamIconUrl": "https://icon",
        },
        "bl1",
        "2025",
    )
    assert match["status"] == "finished"
    assert match["home_score"] == 3
    assert match["matchday"] == 2
    assert team["name"] == "Bayern"
    assert team["short_name"] == "FCB"


def test_teamrankings_worldcup_parsers_normalize_probability_tables():
    crawler = TeamRankingsCrawler()
    advancement_html = """
    <table>
      <thead><tr><th>#</th><th>Team</th><th>Grp</th><th>R32</th><th>R16</th><th>QF</th><th>SF</th><th>Final</th><th>Champ</th></tr></thead>
      <tbody><tr><td>1</td><td>Brazil</td><td>A</td><td>100%</td><td>75%</td><td>50%</td><td>30%</td><td>18%</td><td>10%</td></tr></tbody>
    </table>
    """
    odds_html = """
    <table>
      <thead><tr><th>#</th><th>Team</th><th>Our Model</th><th>Consensus</th><th>Sportsbook</th><th>Polymarket</th><th>Diff</th></tr></thead>
      <tbody><tr><td>1</td><td>Brazil</td><td>14.2%</td><td>12.5%</td><td>13.0%</td><td>12.8%</td><td>+1.2</td></tr></tbody>
    </table>
    """

    class DummyResponse:
        def __init__(self, text):
            self.text = text

    calls = []

    def fake_fetch(url, retries=3, **kwargs):
        calls.append(url)
        return DummyResponse(advancement_html if url.endswith('/advancement') else odds_html)

    crawler._fetch = fake_fetch
    rankings = crawler.crawl("rankings", league="worldcup")
    ratings = crawler.crawl("ratings", league="worldcup")
    assert rankings[0]["team"] == "Brazil"
    assert rankings[0]["champion_pct"] == 10.0
    assert ratings[0]["model_pct"] == 14.2
    assert ratings[0]["market_diff_pct"] == 1.2
    assert any(url.endswith('/advancement') for url in calls)
    assert any(url.endswith('/odds') for url in calls)


def test_excel_exporter_survives_hdfs_snapshot_failure(tmp_path, monkeypatch):
    seed_demo_data()
    from app.export.excel_exporter import ExcelExporter

    def fake_bundle_failure():
        raise RuntimeError("hdfs unavailable")

    monkeypatch.setattr("app.export.excel_exporter.build_group_stage_bundle", fake_bundle_failure)

    db = TestingSessionLocal()
    try:
        exporter = ExcelExporter(db=db, export_dir=str(tmp_path))
        output_path = exporter.export_all()
    finally:
        db.close()

    workbook_path = Path(output_path)
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    assert "分析结果汇总" in workbook.sheetnames
    assert "【清洗后】干净数据" in workbook.sheetnames
    sheet = workbook["【清洗后】干净数据"]
    values = [cell.value for row in sheet.iter_rows(min_row=1, max_row=8) for cell in row if cell.value]
    joined = " ".join(str(value) for value in values)
    assert "失败" in joined or "hdfs unavailable" in joined


def test_team_service_prefers_team_stat_season_over_global_latest():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add(Season(id=6, league_id=3, name="2030"))
        db.commit()

        payload = TeamService().get_team_stats(db, 10)

        assert payload is not None
        assert payload["season"] == "2026"
        assert payload["stats"] is not None
        assert payload["stats"]["goals_for"] == 6
    finally:
        db.close()


def test_player_service_matches_fixed_player_api_default_season_semantics():
    seed_demo_data()
    from app.services.player_service import PlayerService

    db = TestingSessionLocal()
    try:
        service = PlayerService()

        stats_payload = service.get_player_stats(db, 100)
        assert stats_payload["season"] == "2026"
        assert stats_payload["stats"]["goals"] == 2

        compare_payload = service.compare_players(db, 100, 101)
        assert compare_payload["season_stats"]["player_a"]["goals"] == 2
        assert compare_payload["season_stats"]["player_b"] is None
        assert compare_payload["position_rank"]["player_a"]["rank"] == 1
        assert compare_payload["position_rank"]["player_b"]["rank"] is None

        top_scorers = service.get_top_scorers(db, limit=5)
        assert len(top_scorers) == 1
        assert top_scorers[0]["player_id"] == 100

        league_players = service.get_players(db, league_id=3, season="2026", limit=10)
        assert len(league_players) == 1
        assert league_players[0]["id"] == 100

        missing_rank = service.get_position_rank(db, 100, season="2099")
        assert missing_rank["rank"] is None
        assert missing_rank["total"] == 0
    finally:
        db.close()


def test_ingest_team_stats_fallback_uses_matching_league_season_not_global_latest():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.add_all(
            [
                League(id=20, name="Other League", country="Other", type="league"),
                Season(id=30, league_id=20, name="2099"),
            ]
        )
        db.commit()

        stats = crawl_api.ingest_team_stats(
            db,
            [
                {
                    "team": "Brazil",
                    "played": 4,
                    "won": 3,
                    "drawn": 1,
                    "lost": 0,
                    "goals_for": 8,
                    "goals_against": 3,
                }
            ],
            source="fbref",
            league_name="世界杯",
            season_name="2026",
        )

        assert stats["failed"] == 0
        row = db.query(TeamStat).filter(TeamStat.team_id == 10, TeamStat.season_id == 4).first()
        assert row is not None
        wrong_row = db.query(TeamStat).filter(TeamStat.team_id == 10, TeamStat.season_id == 30).first()
        assert wrong_row is None
    finally:
        db.close()


def test_builtin_data_sources_normalize_existing_metadata():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        source.name = "Old FIFA"
        source.type = "api"
        source.priority = 99
        source.description = "stale"
        db.commit()

        changed = ensure_builtin_data_sources(db)
        assert changed >= 1

        refreshed = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert refreshed is not None
        assert refreshed.name == "FIFA Official"
        assert refreshed.type == "crawler"
        assert refreshed.priority != 99
        assert refreshed.description != "stale"
    finally:
        db.close()


def test_builtin_data_sources_clear_orphan_warning_without_logs():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        db.query(CrawlLog).filter(CrawlLog.source_id == source.id).delete()
        source.status = "warning"
        source.error_count = 1
        source.last_crawl_at = datetime.now()
        db.commit()

        changed = ensure_builtin_data_sources(db)
        assert changed >= 1

        refreshed = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert refreshed is not None
        assert refreshed.status == "idle"
        assert refreshed.error_count == 0
        assert refreshed.last_crawl_at is None
    finally:
        db.close()


def test_builtin_data_sources_sync_latest_partial_log_into_source_status():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        source.last_crawl_at = None
        db.add(
            CrawlLog(
                source_id=source.id,
                target="players",
                start_time=datetime(2026, 6, 29, 6, 0, 0),
                end_time=datetime(2026, 6, 29, 6, 0, 5),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()

        changed = ensure_builtin_data_sources(db)
        assert changed >= 0

        refreshed = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert refreshed is not None
        assert refreshed.status == "idle"
        assert refreshed.last_crawl_at is None
    finally:
        db.close()


def test_data_source_health_endpoint_derives_builtin_source_state_without_persisting():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        source.last_crawl_at = None
        db.add(
            CrawlLog(
                source_id=source.id,
                target="players",
                start_time=datetime(2026, 6, 29, 6, 0, 0),
                end_time=datetime(2026, 6, 29, 6, 0, 5),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/health")
    assert response.status_code == 200
    fifa = next(item for item in response.json() if item["source_code"] == "fifa_official")
    assert fifa["status"] == "idle"
    assert fifa["health"] == "idle"
    assert fifa["last_crawl_at"] == "2026-06-29T06:00:05"
    assert fifa["last_log"]["status"] == "partial"


def test_data_source_health_legacy_empty_success_stays_idle_only_after_explicit_reset():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        source.status = "idle"
        source.error_count = 0
        source.last_crawl_at = None
        db.add(
            CrawlLog(
                source_id=source.id,
                target="players",
                start_time=datetime(2026, 6, 29, 6, 0, 0),
                end_time=datetime(2026, 6, 29, 6, 0, 5),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/health")
    assert response.status_code == 200
    fifa = next(item for item in response.json() if item["source_code"] == "fifa_official")
    assert fifa["status"] == "idle"
    assert fifa["health"] == "idle"
    assert fifa["last_log"]["status"] == "partial"
    assert fifa["last_log"]["raw_status"] == "success"


def test_data_source_health_ignores_misaligned_empty_success_logs():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        ensure_builtin_data_sources(db)
        source = db.query(DataSource).filter(DataSource.source_code == "fbref").first()
        assert source is not None
        source.status = "warning"
        source.error_count = 0
        source.last_crawl_at = datetime(2026, 6, 29, 6, 1, 50)
        db.add(
            CrawlLog(
                source_id=source.id,
                target="players",
                start_time=datetime(2026, 6, 29, 6, 1, 50),
                end_time=datetime(2026, 6, 29, 6, 2, 15),
                fetched=0,
                updated=0,
                failed=0,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/health")
    assert response.status_code == 200
    fbref = next(item for item in response.json() if item["source_code"] == "fbref")
    assert fbref["status"] == "idle"
    assert fbref["health"] == "idle"
    assert fbref["last_crawl_at"] is None
    assert fbref["last_log"] is None


def test_data_source_health_ignores_instant_noop_empty_log_when_real_history_exists():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        ensure_builtin_data_sources(db)
        source = db.query(DataSource).filter(DataSource.source_code == "football_data").first()
        assert source is not None
        source.status = "warning"
        source.error_count = 0
        source.last_crawl_at = datetime(2026, 6, 29, 6, 2, 15)
        db.add(
            CrawlLog(
                source_id=source.id,
                target="schedule",
                start_time=datetime(2026, 6, 29, 6, 2, 15),
                end_time=datetime(2026, 6, 29, 6, 2, 15),
                fetched=0,
                updated=0,
                failed=0,
                cost_ms=3,
                status="success",
            )
        )
        db.add(
            CrawlLog(
                source_id=source.id,
                target="matches",
                start_time=datetime(2026, 6, 28, 6, 0, 0),
                end_time=datetime(2026, 6, 28, 6, 0, 30),
                fetched=12,
                updated=12,
                failed=0,
                cost_ms=30000,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/health")
    assert response.status_code == 200
    row = next(item for item in response.json() if item["source_code"] == "football_data")
    assert row["status"] == "active"
    assert row["health"] == "healthy"
    assert row["last_log"]["target"] == "matches"
    assert row["last_log"]["status"] == "success"


def test_data_source_health_keeps_warning_for_only_empty_success_history():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        source = db.query(DataSource).filter(DataSource.source_code == "fifa_official").first()
        assert source is not None
        source.status = "warning"
        source.error_count = 0
        source.last_crawl_at = datetime(2026, 6, 29, 6, 0, 0)
        db.query(CrawlLog).filter(CrawlLog.source_id == source.id).delete()
        db.add(
            CrawlLog(
                source_id=source.id,
                target="players",
                start_time=datetime(2026, 6, 29, 6, 0, 0),
                end_time=datetime(2026, 6, 29, 6, 0, 0),
                fetched=0,
                updated=0,
                failed=0,
                cost_ms=3,
                status="success",
            )
        )
        db.commit()
    finally:
        db.close()

    client = TestClient(app)
    response = client.get("/api/v1/data-sources/health")
    assert response.status_code == 200
    row = next(item for item in response.json() if item["source_code"] == "fifa_official")
    assert row["status"] == "warning"
    assert row["health"] == "warning"
    assert row["last_log"]["target"] == "players"
    assert row["last_log"]["status"] == "partial"


def test_app_lifespan_skips_scheduler_when_disabled(monkeypatch):
    monkeypatch.setattr("app.main.ENABLE_SCHEDULER", False)

    start_calls: list[str] = []
    stop_calls: list[str] = []

    monkeypatch.setattr("app.scheduler.jobs.start_scheduler", lambda: start_calls.append("start"))
    monkeypatch.setattr("app.scheduler.jobs.shutdown_scheduler", lambda: stop_calls.append("stop"))

    async def _run():
        from app.main import lifespan, app as fastapi_app

        async with lifespan(fastapi_app):
            return None

    asyncio.run(_run())

    assert start_calls == []
    assert stop_calls == []


def test_cleanup_stale_crawl_logs_marks_old_running_rows_failed(monkeypatch):
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        log = db.query(CrawlLog).filter(CrawlLog.id == 1).first()
        assert log is not None
        log.status = "running"
        log.end_time = None
        log.start_time = datetime(2026, 6, 28, 0, 0, 0)
        log.error_msg = None
        db.commit()
    finally:
        db.close()

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    import scripts.cleanup_stale_crawl_logs as cleanup_script

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 29, 0, 0, 0)

    monkeypatch.setattr(cleanup_script, "datetime", FrozenDateTime)
    result = cleanup_script.run_cleanup(session_factory=TestingSessionLocal)
    assert result["updated"] == 1
    assert result["updated_ids"] == [1]

    db = TestingSessionLocal()
    try:
        log = db.query(CrawlLog).filter(CrawlLog.id == 1).first()
        assert log is not None
        assert log.status == "failed"
        assert log.end_time == datetime(2026, 6, 29, 0, 0, 0)
        assert "stale running log auto-closed" in (log.error_msg or "")
    finally:
        db.close()


def test_fifa_statistics_aggregates_nonzero_team_metrics():
    crawler = FIFAOfficialCrawler()

    standings = [
        {"team": "Brazil", "played": 1, "won": 1, "drawn": 0, "lost": 0, "goals_for": 2, "goals_against": 0, "league": "世界杯", "season": "2026"},
        {"team": "Argentina", "played": 1, "won": 0, "drawn": 0, "lost": 1, "goals_for": 0, "goals_against": 2, "league": "世界杯", "season": "2026"},
    ]
    matches = [
        {
            "match_id": "m1",
            "fdh_match_id": "fdh-1",
            "home_team": "Brazil",
            "away_team": "Argentina",
            "home_team_id": "10",
            "away_team_id": "11",
            "home_score": 2,
            "away_score": 0,
            "status": "finished",
            "stage": "First Stage",
            "group": "Group A",
        }
    ]
    squad_by_player_id = {
        "100": {"team": "Brazil"},
        "101": {"team": "Brazil"},
        "200": {"team": "Argentina"},
        "201": {"team": "Argentina"},
    }

    crawler._fetch_fdh_player_stats = lambda fdh_match_id, match_id=None: {
        "100": [["XG", 0.7], ["AttemptAtGoal", 4], ["AttemptAtGoalOnTarget", 2], ["Passes", 30], ["PassesCompleted", 24]],
        "101": [["XG", 0.5], ["AttemptAtGoal", 2], ["AttemptAtGoalOnTarget", 1], ["Passes", 20], ["PassesCompleted", 18]],
        "200": [["XG", 0.3], ["AttemptAtGoal", 3], ["AttemptAtGoalOnTarget", 1], ["Passes", 25], ["PassesCompleted", 20]],
        "201": [["XG", 0.2], ["AttemptAtGoal", 1], ["AttemptAtGoalOnTarget", 0], ["Passes", 15], ["PassesCompleted", 12]],
    }

    rows = crawler._build_team_statistics(standings, matches, squad_by_player_id)
    by_team = {row["team"]: row for row in rows}

    assert by_team["Brazil"]["xg_for"] == 1.2
    assert by_team["Brazil"]["xg_against"] == 0.5
    assert by_team["Brazil"]["shots_total"] == 6
    assert by_team["Brazil"]["shots_on_target_total"] == 3
    assert by_team["Brazil"]["passes_total"] == 50
    assert by_team["Brazil"]["pass_accuracy"] == 84.0
    assert by_team["Brazil"]["clean_sheets"] == 1

    assert by_team["Argentina"]["xg_for"] == 0.5
    assert by_team["Argentina"]["xg_against"] == 1.2
    assert by_team["Argentina"]["passes_total"] == 40


def test_worldcup_coverage_reports_partial_when_only_player_xg_exists():
    seed_demo_data()
    db = TestingSessionLocal()
    try:
        db.query(Shot).delete()
        stat = db.query(TeamStat).filter(TeamStat.id == 1).first()
        assert stat is not None
        stat.shots_total = 0
        stat.shots_on_target_total = 0
        db.commit()

        from app.api.worldcup import get_worldcup_coverage

        payload = get_worldcup_coverage("2026", db)
        shots_module = next(item for item in payload["coverage"] if item["module"] == "射门热图 / xG 时间线")
        assert shots_module["status"] == "partial"
        assert "shots_table=0" in shots_module["detail"]
    finally:
        db.close()
